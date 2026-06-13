"""End-to-End-Smoketest gegen echte JA-PDFs.

Pipeline: Read PDF -> classify -> extract via Claude -> consolidate ->
build_excel. Speichert das Excel und reportet PDF-JUE vs Excel-JUE-Diff.

Usage:
    .venv/bin/python tests/fixtures/smoketest_e2e.py "<pfad/zur/ja.pdf>" [<weitere ja.pdfs>...]
"""
import io
import sys
import time
from pathlib import Path
from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parents[2]
load_dotenv(ROOT / ".env")
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402

from app.worker.claude_client import ClaudeClient  # noqa: E402
from app.worker.consolidate import merge_extractions  # noqa: E402
from app.worker.pdf_detect import (  # noqa: E402
    classify_pdf, extract_guv_section, pdf_to_images, PdfKind,
)
from app.excel.builder import build_excel  # noqa: E402


def extract_one(client: ClaudeClient, pdf_path: Path) -> dict:
    print(f"\n→ {pdf_path.name}")
    data = pdf_path.read_bytes()
    kind = classify_pdf(data)
    print(f"  kind={kind.value}, size={len(data):,}B")
    from app.worker.pdf_detect import extract_text as _extract_text
    t0 = time.time()
    if kind == PdfKind.TEXT:
        # doc_type detection — Susa & BWA brauchen anderen Prompt-Pfad als JA
        full_text = _extract_text(data)
        doc_type = client.classify_document(full_text[:5000])
        print(f"  doc_type={doc_type}")
        if doc_type in ("bwa", "susa"):
            text = full_text  # ganzer Text — Susa/BWA sind kurz
        else:
            text = extract_guv_section(data)
            doc_type = "jahresabschluss"
        print(f"  feeding {len(text):,} chars to Claude as '{doc_type}'")
        result = client.extract_text_pdf(text, doc_type=doc_type)
    else:
        doc_type = client.classify_document("SCAN-BILDDATEN")
        print(f"  doc_type={doc_type}")
        images = pdf_to_images(data, dpi=100)
        print(f"  scan with {len(images)} pages")
        result = client.extract_scan_pdf(images, doc_type=doc_type)
    dt = time.time() - t0
    print(f"  Claude returned in {dt:.1f}s")
    return result


def report(name: str, result: dict) -> None:
    n_groups = len(result.get("groups", []))
    n_acc = sum(len(g.get("accounts", [])) for g in result.get("groups", []))
    # gkv_section kann bei EÜR-spezifischen Gruppen None sein → str-coerce.
    sections = sorted({(g.get("gkv_section") or "<none>")
                       for g in result.get("groups", [])})
    print(f"  type={result.get('type')}, year={result.get('year')}, "
          f"prev={result.get('previous_year')}, sign={result.get('sign_convention')}")
    print(f"  endwert_label={result.get('endwert_label')!r}")
    print(f"  groups={n_groups}, accounts={n_acc}")
    print(f"  pdf_jue_gj={result.get('pdf_jahresueberschuss_gj')}, "
          f"pdf_jue_vj={result.get('pdf_jahresueberschuss_vj')}")
    print(f"  gkv_sections seen: {sections}")


def main(pdf_paths: list[Path]) -> int:
    from app.worker.tasks import _extract_pdf  # gleiche Logik wie Produktion
    client = ClaudeClient()
    extractions = []
    for p in pdf_paths:
        if not p.exists():
            print(f"FEHLER: {p} nicht gefunden", file=sys.stderr)
            return 2
        print(f"\n→ {p.name}")
        results = _extract_pdf(client, p.read_bytes())
        if len(results) > 1:
            print(f"  → Bundle: {len(results)} Extraktionen "
                  f"({', '.join(r.get('type') for r in results)})")
        for result in results:
            result["file"] = p.name
            report(p.name, result)
            extractions.append(result)

    print("\n→ consolidate")
    cons = merge_extractions(extractions)
    print(f"  columns: {[c['label'] for c in cons['columns']]}")
    print(f"  groups: {len(cons['groups'])}")
    print(f"  questions: {len(cons['questions'])}")
    print(f"  pdf_jue_per_column: {cons.get('pdf_jue_per_column')}")

    print("\n→ build_excel")
    xlsx = build_excel(cons)

    out_path = ROOT / "smoketest_output.xlsx"
    out_path.write_bytes(xlsx)
    print(f"  wrote {out_path} ({len(xlsx):,} bytes)")

    # Cross-Check: lade die Excel und vergleiche Endwert-Formel vs PDF-Endwert.
    # Bei EÜR heisst die Zeile 'Steuerlicher Gewinn ...', bei HGB 'Jahresergebnis'.
    wb = load_workbook(io.BytesIO(xlsx), data_only=False)
    ws = wb["Übertrag"]

    print("\n→ Plausibilitaets-Cross-Check")
    expected_label = cons.get("endwert_label") or "Jahresergebnis"
    je_row = None
    for row in ws.iter_rows():
        if row[1].value == expected_label:
            je_row = row[0].row
            break
    if je_row is None:
        print(f"  WARN: keine Endwert-Zeile gefunden (gesucht: {expected_label!r})")
        return 1
    print(f"  Endwert-Zeile '{expected_label}' (Row {je_row}), Formel pro Spalte:")
    for col_idx, col in enumerate(cons["columns"]):
        formula = ws.cell(je_row, 3 + col_idx).value
        print(f"    {col['label']}: {formula}")

    # Fragen-Sheet zeigen (falls vorhanden)
    if "Fragen" in wb.sheetnames:
        fragen = wb["Fragen"]
        print(f"\n→ Fragen-Sheet ({fragen.max_row - 1} Eintraege)")
        for row in fragen.iter_rows(min_row=2, values_only=True):
            if row[0]:
                print(f"  - {row[0]}: {row[1]}")
    else:
        print("\n→ Fragen-Sheet: keins (alles sauber)")

    return 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    paths = [Path(a) for a in sys.argv[1:]]
    sys.exit(main(paths))
