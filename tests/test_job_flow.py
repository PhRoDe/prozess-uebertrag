"""End-to-End-Test des gesamten Flows mit In-Memory-Fakes für Supabase
und gemocktem Claude. Beweist, dass alle Module zusammenspielen.

Live-Test mit echter API läuft separat via tests/fixtures/smoketest_claude.py.
"""
import io
from datetime import datetime, timedelta, timezone
from unittest.mock import patch

import fitz
from openpyxl import load_workbook
from fastapi.testclient import TestClient

from app.main import app
from app.models import InputFile, Job, JobStatus


def _minimal_ja_pdf() -> bytes:
    """Create a text-PDF that looks like a Kontennachweis."""
    doc = fitz.open()
    page = doc.new_page()
    lines = [
        "Jahresabschluss 2024",
        "Kontennachweis GuV",
        "1. Umsatzerloese",
        "  8400  Erloese 19% USt  1.000.000,00  900.000,00",
        "4. Materialaufwand",
        "  5100  Wareneingang  400.000,00  370.000,00",
    ]
    for i, line in enumerate(lines):
        page.insert_text((72, 72 + i * 14), line)
    buf = io.BytesIO()
    doc.save(buf)
    doc.close()
    return buf.getvalue()


CLAUDE_EXTRACTION = {
    "type": "jahresabschluss", "year": 2024, "previous_year": 2023,
    "accounts": [
        {"konto_nr": "8400", "bezeichnung": "Erloese 19%",
         "gruppe": "1. Umsatzerlöse",
         "betrag_gj": 1000000.0, "betrag_vj": 900000.0, "confidence": "high"},
        {"konto_nr": "5100", "bezeichnung": "Wareneingang",
         "gruppe": "4a. Aufwendungen für RHB und Waren",
         "betrag_gj": 400000.0, "betrag_vj": 370000.0, "confidence": "high"},
    ],
    "open_questions": [],
}


class FakeRepo:
    def __init__(self):
        self.jobs: dict[str, Job] = {}
        self.client = type("C", (), {
            "table": lambda self, n: type("T", (), {
                "delete": lambda s: type("D", (), {"eq": lambda s, k, v: type("E", (), {"execute": lambda s: None})()})(),
            })(),
        })()

    def create(self, input_files):
        now = datetime.now(timezone.utc)
        job = Job(
            id=f"job-{len(self.jobs) + 1}", created_at=now,
            status=JobStatus.UPLOADED, input_files=input_files,
            expires_at=now + timedelta(hours=24),
        )
        self.jobs[job.id] = job
        return job

    def get(self, job_id):
        return self.jobs.get(job_id)

    def set_status(self, job_id, status, error=None):
        if job_id in self.jobs:
            self.jobs[job_id].status = status
            if error:
                self.jobs[job_id].error_message = error

    def set_input_files(self, job_id, files):
        self.jobs[job_id].input_files = files

    def set_extraction(self, job_id, extraction):
        self.jobs[job_id].extraction = extraction
        self.jobs[job_id].status = JobStatus.REVIEW_NEEDED

    def set_output(self, job_id, path, answers):
        self.jobs[job_id].output_path = path
        self.jobs[job_id].review_answers = answers
        self.jobs[job_id].status = JobStatus.READY

    def try_claim(self, job_id, node_id):
        return True

    def release_claim(self, job_id):
        pass

    def list_resumable(self):
        return []


class FakeStorage:
    def __init__(self):
        self.store: dict[str, bytes] = {}

    def upload_input(self, job_id, filename, data):
        path = f"{job_id}/input/{filename}"
        self.store[path] = data
        return path

    def upload_output(self, job_id, data):
        path = f"{job_id}/output.xlsx"
        self.store[path] = data
        return path

    def download_input(self, path):
        return self.store[path]

    def signed_output_url(self, path, expires_in=900):
        return f"https://fake.supabase.co/signed/{path}?token=x"

    def delete_job(self, job_id):
        for p in list(self.store):
            if p.startswith(f"{job_id}/"):
                del self.store[p]


def test_full_flow_upload_extract_finalize_download():
    fake_repo = FakeRepo()
    fake_storage = FakeStorage()

    with patch("app.routes.upload.JobsRepo", return_value=fake_repo), \
         patch("app.routes.upload.StorageClient", return_value=fake_storage), \
         patch("app.routes.job.JobsRepo", return_value=fake_repo), \
         patch("app.routes.download.JobsRepo", return_value=fake_repo), \
         patch("app.routes.download.StorageClient", return_value=fake_storage), \
         patch("app.worker.tasks.JobsRepo", return_value=fake_repo), \
         patch("app.worker.tasks.StorageClient", return_value=fake_storage), \
         patch("app.worker.tasks.ClaudeClient") as ClaudeMock, \
         patch("app.routes.pages.validate_session_token", return_value=True):

        ClaudeMock.return_value.classify_document.return_value = "jahresabschluss"
        ClaudeMock.return_value.extract_text_pdf.return_value = CLAUDE_EXTRACTION

        client = TestClient(app)
        pdf_bytes = _minimal_ja_pdf()

        # Upload (cookie set to pass require_auth via patched validator)
        r = client.post(
            "/upload",
            files=[("files", ("ja-2024.pdf", pdf_bytes, "application/pdf"))],
            cookies={"pu_session": "fake-valid"},
        )
        assert r.status_code == 200, r.text
        redirect = r.headers.get("HX-Redirect")
        assert redirect and redirect.startswith("/job/")
        job_id = redirect.split("/")[-1]

        # TestClient executes BackgroundTasks synchronously after response
        assert fake_repo.jobs[job_id].status == JobStatus.REVIEW_NEEDED

        # Finalize
        from app.worker.tasks import finalize_job
        finalize_job(job_id, {})
        assert fake_repo.jobs[job_id].status == JobStatus.READY

        # Excel contains real content
        xlsx_bytes = fake_storage.store[f"{job_id}/output.xlsx"]
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert "Übertrag" in wb.sheetnames
        ws = wb["Übertrag"]
        has_sum_formula = any(
            isinstance(c.value, str) and c.value.startswith("=SUM")
            for row in ws.iter_rows() for c in row
        )
        assert has_sum_formula, "Excel contains no SUM formulas"

        # Download endpoint redirects to signed URL
        r = client.get(f"/download/{job_id}", cookies={"pu_session": "fake-valid"},
                       follow_redirects=False)
        assert r.status_code == 303
        assert "fake.supabase.co/signed/" in r.headers["Location"]
