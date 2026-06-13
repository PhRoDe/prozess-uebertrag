import io
import json
import pytest
from openpyxl import load_workbook
from app.excel.builder import build_excel, _inject_restposten_accounts


def _sample(sign="expenses_negative", with_bwa=False):
    cols = [
        {"label": "2023", "kind": "ja", "year": 2023, "sign_convention": sign},
        {"label": "2024", "kind": "ja", "year": 2024, "sign_convention": sign},
    ]
    if with_bwa:
        cols.append({"label": "BWA 2025", "kind": "bwa", "year": 2025,
                     "sign_convention": sign})
    groups = [
        {"name": "Umsatzerlöse", "type": "ertrag", "sub_group_of": None,
         "column_sums": {},
         "accounts": [
             {"konto_nr": "8400", "bezeichnung": "Erlöse 19%",
              "values": {0: 900000, 1: 1000000}, "confidence": "high"},
         ]},
        {"name": "Materialaufwand", "type": "aufwand", "sub_group_of": None,
         "column_sums": {},
         "accounts": [
             {"konto_nr": "5100", "bezeichnung": "Wareneingang",
              "values": {0: -370000, 1: -400000} if sign == "expenses_negative"
                        else {0: 370000, 1: 400000},
              "confidence": "high"},
         ]},
    ]
    if with_bwa:
        groups[0]["column_sums"][2] = 500000
        groups[1]["column_sums"][2] = -200000 if sign == "expenses_negative" else 200000
    return {"columns": cols, "groups": groups, "questions": []}


def _ws(xlsx):
    return load_workbook(io.BytesIO(xlsx))["Übertrag"]


def _find_row(ws, label):
    for row in ws.iter_rows():
        if row[1].value == label:
            return row[0].row
    return None


def test_header_has_konto_bezeichnung_and_year_columns():
    xlsx = build_excel(_sample())
    ws = _ws(xlsx)
    assert ws.cell(1, 1).value == "Konto"
    assert ws.cell(1, 2).value == "Bezeichnung"
    assert ws.cell(1, 3).value == "2023"
    assert ws.cell(1, 4).value == "2024"


def test_group_sum_row_comes_before_details():
    xlsx = build_excel(_sample())
    ws = _ws(xlsx)
    umsatz_sum = _find_row(ws, "Umsatzerlöse")
    erloese_detail = _find_row(ws, "  Erlöse 19%")
    assert umsatz_sum is not None and erloese_detail is not None
    assert umsatz_sum < erloese_detail


def test_ja_sums_are_formulas():
    xlsx = build_excel(_sample())
    ws = _ws(xlsx)
    umsatz_row = _find_row(ws, "Umsatzerlöse")
    val = ws.cell(umsatz_row, 3).value  # 2023-Spalte
    assert isinstance(val, str) and val.startswith("=SUM")


def test_bwa_sum_when_acc_sum_differs_from_pdf_sum_via_restposten():
    """Wenn die Gruppe Konten hat aber column_sum (= pdf_sum_gj) fuer die
    Spalte abweicht: Restposten-Konto ergaenzt, Excel-Gruppen-Zelle bleibt
    `=SUM(...)`-Formel. Excel-Wert == pdf_sum_gj durch den Restposten.

    CLAUDE.md-Regel: "Alle Excel-Zwischensummen MÜSSEN Formeln sein (=SUM(...)
    oder Kaskaden). Niemals hardcoded Werte." → also Restposten als zusätzliche
    Detail-Zeile, nicht column_sum-Direktwert in der Sum-Zelle.
    """
    xlsx = build_excel(_sample(with_bwa=True))
    ws = _ws(xlsx)
    umsatz_row = _find_row(ws, "Umsatzerlöse")
    # Alle Spalten haben SUM-Formel (auch die BWA-Spalte mit Restposten)
    for col in (3, 4, 5):
        val = ws.cell(umsatz_row, col).value
        assert isinstance(val, str) and val.startswith("=SUM"), \
            f"Spalte {col} sollte SUM-Formel haben, ist {val!r}"
    # Restposten-Zeile existiert
    restposten_row = _find_row(ws, "  Restposten — nicht aufgeschlüsselt im PDF")
    assert restposten_row is not None, "Restposten-Konto sollte ergänzt sein"
    # Restposten 2023 (Spalte 3) = pdf_sum_gj - acc_sum, hier 0 weil cs gleich acc
    # Restposten 2025 (Spalte 5, BWA) = 500000 - 0 = 500000
    assert ws.cell(restposten_row, 5).value == 500000


def test_inject_restposten_skips_bestandsveraenderung():
    """Regression Prisma 2026-06: Bestandsveraenderung darf NIE einen Restposten
    bekommen.

    Das Detail-Konto ist vorzeichen-normalisiert (Verminderung -> negativ,
    `_normalize_bestand_value`), der Plausibilitaets-Anker `column_sums`
    (= pdf_sum_gj) bleibt aber roh positiv wie im PDF gedruckt. Ein Restposten
    = Anker - Detailsumme = (+1,64M) - (-1,64M) = +3,28M wuerde das Vorzeichen
    der Position kippen -> JUE-Fehler in Millionenhoehe (real: JA2023 +3.283.620,46,
    JA2024 +1.848.360,50). Eine Bestandsveraenderung ist ein einzelner
    vorzeichenbehafteter Delta-Wert, kein 'fehlende Konten'-Fall -> kein Restposten.
    """
    groups = [{
        "name": "2. Verminderung des Bestandes an fertigen und unfertigen Erzeugnissen",
        "type": "ertrag", "gkv_section": "bestandsveraenderung", "sub_group_of": None,
        "column_sums": {0: 1641810.23},  # roher Anker, positiv wie im PDF
        "accounts": [{"konto_nr": "", "bezeichnung": "Bestandsveränderung Bauaufträge",
                      "values": {0: -1641810.23}, "confidence": "high"}],  # normalisiert negativ
    }]
    out = _inject_restposten_accounts(groups)
    accs = out[0]["accounts"]
    assert len(accs) == 1, "Bestandsveraenderung darf kein Restposten-Konto bekommen"
    assert all("Restposten" not in (a.get("bezeichnung") or "") for a in accs)


def test_inject_restposten_still_fills_non_bestand_groups():
    """Guard: der Bestand-Skip darf den regulaeren Restposten-Mechanismus
    (Bilanzbericht: acc_sum < pdf_sum_gj) NICHT abschalten."""
    groups = [{
        "name": "7. sonstige betriebliche Aufwendungen", "type": "aufwand",
        "gkv_section": "sonst_betr_aufwand", "sub_group_of": None,
        "column_sums": {0: -100000.0},
        "accounts": [{"konto_nr": "6800", "bezeichnung": "Bueromaterial",
                      "values": {0: -65000.0}, "confidence": "high"}],
    }]
    out = _inject_restposten_accounts(groups)
    accs = out[0]["accounts"]
    assert len(accs) == 2
    rest = accs[-1]
    assert "Restposten" in rest["bezeichnung"]
    assert rest["values"][0] == -35000.0  # -100000 - (-65000)


def test_build_excel_bestandsveraenderung_verminderung_no_phantom_restposten():
    """End-to-End über den ganzen Builder (Prisma 2026-06): eine Verminderungs-
    Bestandsgruppe mit normalisiert-negativem Detail (-1.641.810,23) und rohem
    positivem Anker (+1.641.810,23) darf KEINE Restposten-Zeile bekommen, und die
    Gruppen-Sum-Zelle muss =SUM über genau das eine Detail-Konto sein (negativ).
    Sonst kippt das Vorzeichen und die JÜ-Formel ist 2×|wert| daneben.
    """
    cons = {
        "columns": [{"label": "2023", "kind": "ja", "year": 2023,
                     "sign_convention": "expenses_positive",
                     "pdf_jahresueberschuss_gj": 501424.69}],
        "groups": [
            {"name": "1. Umsatzerlöse", "type": "ertrag", "sub_group_of": None,
             "column_sums": {}, "accounts": [
                {"konto_nr": "8400", "bezeichnung": "Erlöse", "values": {0: 5523319.58},
                 "confidence": "high"}]},
            {"name": "2. Verminderung des Bestandes an fertigen und unfertigen Erzeugnissen",
             "type": "ertrag", "gkv_section": "bestandsveraenderung", "sub_group_of": None,
             "column_sums": {0: 1641810.23}, "accounts": [
                {"konto_nr": "", "bezeichnung": "Bestandsveränderung Bauaufträge",
                 "values": {0: -1641810.23}, "confidence": "high"}]},
        ],
        "questions": [],
    }
    ws = _ws(build_excel(cons))
    assert _find_row(ws, "  Restposten — nicht aufgeschlüsselt im PDF") is None, \
        "Bestandsveränderung darf keine Restposten-Zeile erzeugen"
    bestand_row = _find_row(
        ws, "2. Verminderung des Bestandes an fertigen und unfertigen Erzeugnissen")
    formula = ws.cell(bestand_row, 3).value
    assert isinstance(formula, str) and formula.startswith("=SUM"), \
        "Gruppen-Summe muss Formel bleiben"
    # Das einzige Detail-Konto direkt unter der Gruppe trägt den negativen Wert
    detail = ws.cell(bestand_row + 1, 3).value
    assert detail == -1641810.23


def test_jahresergebnis_expenses_negative_is_simple_sum():
    """expenses_negative: alle Gruppen addieren, weil Aufwände eh negativ sind."""
    xlsx = build_excel(_sample(sign="expenses_negative"))
    ws = _ws(xlsx)
    je_row = _find_row(ws, "Jahresergebnis")
    formula = ws.cell(je_row, 3).value
    assert isinstance(formula, str) and formula.startswith("=")
    # Nur Pluszeichen, keine Minus-Zeichen
    assert "-" not in formula.replace("=", "")


def test_jahresergebnis_expenses_positive_uses_minus_for_aufwand():
    """expenses_positive: Aufwände müssen explizit subtrahiert werden."""
    xlsx = build_excel(_sample(sign="expenses_positive"))
    ws = _ws(xlsx)
    je_row = _find_row(ws, "Jahresergebnis")
    formula = ws.cell(je_row, 3).value
    assert isinstance(formula, str) and formula.startswith("=")
    # Muss ein Minus enthalten (Materialaufwand wird abgezogen)
    assert "-" in formula.replace("=", "")


def test_fragen_sheet_lists_questions():
    cons = _sample()
    cons["questions"] = [
        {"type": "previous_year_mismatch", "group": "Umsatzerlöse",
         "konto_nr": "8400", "year": 2023, "from_doc_year": 2024,
         "pdf_says": 1000000, "own_value": 999000},
    ]
    xlsx = build_excel(cons)
    wb = load_workbook(io.BytesIO(xlsx))
    assert "Fragen" in wb.sheetnames
    fragen = wb["Fragen"]
    rows = list(fragen.iter_rows(values_only=True))
    assert rows[0] == ("Thema", "Details")
    assert rows[1][0] == "previous_year_mismatch"


def test_pdf_jue_row_rendered_when_pdf_value_present():
    """Wenn pdf_jue_per_column geliefert wird, rendert der Builder eine
    'PDF-Jahresüberschuss'-Zeile mit den Werten und eine
    'Differenz Excel ↔ PDF'-Zeile mit Subtraktions-Formel."""
    cons = _sample()
    cons["pdf_jue_per_column"] = {0: 530000.00, 1: 600000.00}
    xlsx = build_excel(cons)
    ws = _ws(xlsx)
    pdf_jue_row = _find_row(ws, "PDF-Jahresüberschuss")
    diff_row = _find_row(ws, "Differenz Excel ↔ PDF")
    assert pdf_jue_row is not None
    assert diff_row is not None
    assert ws.cell(pdf_jue_row, 3).value == 530000.00
    assert ws.cell(pdf_jue_row, 4).value == 600000.00
    diff_formula = ws.cell(diff_row, 3).value
    assert isinstance(diff_formula, str) and diff_formula.startswith("=")


def test_pdf_jue_mismatch_creates_fragen_entry():
    """JÜ-Formel weicht vom PDF-Wert ab → Excel wird trotzdem gebaut, aber
    die Diff erscheint im Fragen-Sheet (sichtbar für den User) und in der
    'Differenz Excel ↔ PDF'-Zeile pro Spalte. Vorher: ValueError → keine
    Excel → User stuck. Jetzt: User bekommt die Excel, sieht die Diff
    prominent in der Anker-Zeile und im Fragen-Sheet, kann manuell prüfen
    (Live-Bug 2026-05-12 Job df31b6cd: Bilanzbericht-Mandat mit
    unvollständig extrahierten Konten — hard-fail blockierte den User
    obwohl die Excel grundsätzlich nutzbar gewesen wäre)."""
    cons = _sample()
    cons["pdf_jue_per_column"] = {0: 530000.00, 1: 700000.00}
    xlsx = build_excel(cons)
    wb = load_workbook(io.BytesIO(xlsx))
    assert "Fragen" in wb.sheetnames, "Fragen-Sheet muss bei JÜ-Mismatch angelegt werden"
    fragen = wb["Fragen"]
    fragen_rows = list(fragen.iter_rows(values_only=True))
    jue_mismatches = [r for r in fragen_rows if r[0] == "jue_excel_vs_pdf_mismatch"]
    assert len(jue_mismatches) >= 1, \
        f"Mindestens ein jue_excel_vs_pdf_mismatch-Eintrag erwartet, gefunden: {fragen_rows}"


def test_no_fragen_sheet_when_clean():
    """Bei sauberem Lauf (keine echten User-Entscheidungen offen) wird das
    Fragen-Sheet gar nicht erst angelegt."""
    cons = _sample()
    cons["questions"] = []
    xlsx = build_excel(cons)
    wb = load_workbook(io.BytesIO(xlsx))
    assert "Fragen" not in wb.sheetnames


def test_bilanzgewinn_block_separate_from_jue():
    """Gewinnvortrag/Ausschüttung/Bilanzgewinn sind NICHT im JÜ enthalten,
    sondern werden als eigener Block nach dem JÜ gerendert. Der Bilanzgewinn
    ist eine Formel (= JÜ + Gewinnvortrag - Ausschüttung)."""
    cons = {
        "columns": [
            {"label": "2024", "kind": "ja", "year": 2024,
             "sign_convention": "expenses_negative"},
        ],
        "groups": [
            {"name": "Umsatzerlöse", "type": "ertrag",
             "gkv_section": "umsatzerloese",
             "sub_group_of": None, "column_sums": {},
             "accounts": [{"konto_nr": "8400", "bezeichnung": "Erlöse",
                           "values": {0: 1000000}, "confidence": "high"}]},
            {"name": "Materialaufwand", "type": "aufwand",
             "gkv_section": "materialaufwand_rhb",
             "sub_group_of": None, "column_sums": {},
             "accounts": [{"konto_nr": "5100", "bezeichnung": "Wareneing.",
                           "values": {0: -400000}, "confidence": "high"}]},
            {"name": "Gewinnvortrag", "type": "neutral",
             "gkv_section": "gewinnvortrag",
             "sub_group_of": None, "column_sums": {},
             "accounts": [{"konto_nr": "8990", "bezeichnung": "Vortrag",
                           "values": {0: 50000}, "confidence": "high"}]},
            {"name": "Ausschüttung", "type": "neutral",
             "gkv_section": "ausschuettung",
             "sub_group_of": None, "column_sums": {},
             "accounts": [{"konto_nr": "8995", "bezeichnung": "Ausschüttung",
                           "values": {0: 100000}, "confidence": "high"}]},
        ],
        "questions": [],
    }
    xlsx = build_excel(cons)
    ws = _ws(xlsx)
    je_row = _find_row(ws, "Jahresergebnis")
    bilanzgewinn_row = _find_row(ws, "Bilanzgewinn (Formel)")
    gewinnvortrag_sum_row = _find_row(ws, "Gewinnvortrag")
    ausschuettung_sum_row = _find_row(ws, "Ausschüttung")
    # Bilanzgewinn-Block kommt NACH dem JÜ
    assert bilanzgewinn_row is not None
    assert bilanzgewinn_row > je_row
    assert gewinnvortrag_sum_row > je_row
    assert ausschuettung_sum_row > je_row
    # JÜ-Formel enthaelt nur die GuV-Gruppen, nicht Gewinnvortrag/Ausschuettung
    je_formula = ws.cell(je_row, 3).value
    assert isinstance(je_formula, str)
    # Zeilen-Refs in Formel duerfen Gewinnvortrag/Ausschuettung NICHT enthalten
    assert f"C{gewinnvortrag_sum_row}" not in je_formula
    assert f"C{ausschuettung_sum_row}" not in je_formula
    # Bilanzgewinn = JÜ + Gewinnvortrag - Ausschuettung
    bg_formula = ws.cell(bilanzgewinn_row, 3).value
    assert isinstance(bg_formula, str) and bg_formula.startswith("=")


def test_jue_uses_gkv_section_for_classification():
    """Wenn Claude eine Aufwand-Gruppe versehentlich als type='neutral'
    klassifiziert, aber gkv_section korrekt ist, wird sie trotzdem
    abgezogen (gkv_section ist authoritativ)."""
    cons = {
        "columns": [
            {"label": "2024", "kind": "ja", "year": 2024,
             "sign_convention": "expenses_positive"},
        ],
        "groups": [
            {"name": "Umsatzerlöse", "type": "ertrag",
             "gkv_section": "umsatzerloese",
             "sub_group_of": None, "column_sums": {},
             "accounts": [{"konto_nr": "8400", "bezeichnung": "Erlöse",
                           "values": {0: 1000000}, "confidence": "high"}]},
            # Personalaufwand mit type=neutral, aber gkv_section=...loehne
            {"name": "Personalaufwand", "type": "neutral",
             "gkv_section": "personalaufwand_loehne",
             "sub_group_of": None, "column_sums": {},
             "accounts": [{"konto_nr": "6000", "bezeichnung": "Löhne",
                           "values": {0: 300000}, "confidence": "high"}]},
        ],
        "questions": [],
    }
    xlsx = build_excel(cons)
    ws = _ws(xlsx)
    je_row = _find_row(ws, "Jahresergebnis")
    formula = ws.cell(je_row, 3).value
    # Personalaufwand muss subtrahiert werden -> Minus in Formel
    assert "-" in formula.replace("=", "")


def test_pdf_jue_row_skipped_when_no_pdf_value():
    """Ohne pdf_jue_per_column wird keine PDF-JÜ-Zeile gerendert."""
    cons = _sample()
    xlsx = build_excel(cons)
    ws = _ws(xlsx)
    assert _find_row(ws, "PDF-Jahresüberschuss") is None


def test_top_level_with_own_accounts_and_subgroups_sums_both():
    """Wenn eine Top-Level-Gruppe sowohl eigene Konten als auch Sub-Gruppen hat
    (zB 'Sonst. betr. Aufw.' mit direktem Konto 'Forderungsverlust' + den
    Sub-Gruppen 'Raumkosten', 'Versicherungen'), muss die Top-Level-Summe
    BEIDES enthalten."""
    cons = {
        "columns": [
            {"label": "2024", "kind": "ja", "year": 2024,
             "sign_convention": "expenses_negative"},
        ],
        "groups": [
            {"name": "Sonst. betr. Aufw.", "type": "aufwand",
             "gkv_section": "sonst_betr_aufw",
             "sub_group_of": None, "column_sums": {},
             "accounts": [{"konto_nr": "6960", "bezeichnung": "Forderungsverlust",
                           "values": {0: -100}, "confidence": "high"}]},
            {"name": "Raumkosten", "type": "aufwand",
             "gkv_section": "sonst_betr_aufw",
             "sub_group_of": "Sonst. betr. Aufw.", "column_sums": {},
             "accounts": [{"konto_nr": "4210", "bezeichnung": "Miete",
                           "values": {0: -1000}, "confidence": "high"}]},
            {"name": "Versicherungen", "type": "aufwand",
             "gkv_section": "sonst_betr_aufw",
             "sub_group_of": "Sonst. betr. Aufw.", "column_sums": {},
             "accounts": [{"konto_nr": "4360", "bezeichnung": "Beitr.",
                           "values": {0: -200}, "confidence": "high"}]},
        ],
        "questions": [],
    }
    xlsx = build_excel(cons)
    ws = _ws(xlsx)
    parent_row = _find_row(ws, "Sonst. betr. Aufw.")
    formula = ws.cell(parent_row, 3).value
    # Formel muss SUM(eigene Konten) UND Sub-Gruppen-Refs enthalten
    assert isinstance(formula, str) and formula.startswith("=")
    raumkosten_row = _find_row(ws, "  Raumkosten")
    versicherungen_row = _find_row(ws, "  Versicherungen")
    assert raumkosten_row is not None and versicherungen_row is not None
    # Beide Sub-Refs muessen in der Top-Level-Formel auftauchen
    assert f"C{raumkosten_row}" in formula
    assert f"C{versicherungen_row}" in formula


def test_subgroup_sum_indented():
    """Sub-Gruppen bekommen Einrückung durch zwei führende Leerzeichen."""
    cons = {
        "columns": [
            {"label": "2024", "kind": "ja", "year": 2024,
             "sign_convention": "expenses_negative"},
        ],
        "groups": [
            {"name": "5. Sonstige betriebliche Aufwendungen", "type": "aufwand",
             "sub_group_of": None, "column_sums": {}, "accounts": []},
            {"name": "5.1 Versicherungen", "type": "aufwand",
             "sub_group_of": "5. Sonstige betriebliche Aufwendungen",
             "column_sums": {}, "accounts": [
                {"konto_nr": "4380", "bezeichnung": "Beiträge",
                 "values": {0: -500}, "confidence": "high"},
             ]},
        ],
        "questions": [],
    }
    xlsx = build_excel(cons)
    ws = _ws(xlsx)
    sub_row = _find_row(ws, "  5.1 Versicherungen")  # mit Einrückung
    assert sub_row is not None


def test_builder_survives_json_roundtrip():
    """Postgres JSONB turns int keys into strings. The builder must still
    produce non-empty value cells after a json.dumps -> json.loads cycle."""
    cons = _sample()
    cons_after_db = json.loads(json.dumps(cons))
    xlsx = build_excel(cons_after_db)
    ws = _ws(xlsx)
    erloese_row = _find_row(ws, "  Erlöse 19%")
    assert erloese_row is not None
    assert ws.cell(erloese_row, 3).value == 900000
    assert ws.cell(erloese_row, 4).value == 1000000


def test_builder_raises_when_all_account_values_empty():
    """Sanity guard: if every account value is None/empty after the build,
    something is structurally wrong (most likely a roundtrip bug). Fail loud
    instead of silently producing a useless Excel."""
    cons = _sample()
    for g in cons["groups"]:
        for acc in g["accounts"]:
            acc["values"] = {}
    with pytest.raises(ValueError, match="empty"):
        build_excel(cons)


def test_jue_formel_addiert_bestandsveraenderung():
    """JÜ-Formel addiert die Bestandsveränderungs-Gruppe (Werte sind im
    consolidated bereits signed: + Erhöhung / − Verminderung). Sie darf
    NICHT als Aufwand subtrahiert werden, sonst würde sie doppelt wirken."""
    cons = {
        "columns": [
            {"label": "2024", "kind": "ja", "year": 2024,
             "sign_convention": "expenses_negative"},
        ],
        "groups": [
            {"name": "Umsatzerlöse", "type": "ertrag",
             "gkv_section": "umsatzerloese", "sub_group_of": None,
             "column_sums": {},
             "accounts": [{"konto_nr": "8400", "bezeichnung": "Erlöse",
                            "values": {0: 1000000}, "confidence": "high"}]},
            {"name": "Verminderung des Bestandes an fertigen und unfertigen Erzeugnissen",
             "type": "ertrag", "gkv_section": "bestandsveraenderung",
             "sub_group_of": None, "column_sums": {},
             "accounts": [{"konto_nr": "4815", "bezeichnung": "Bestandsv.",
                            "values": {0: -614000}, "confidence": "high"}]},
        ],
        "questions": [],
    }
    xlsx = build_excel(cons)
    ws = _ws(xlsx)
    je_row = _find_row(ws, "Jahresergebnis")
    bestand_row = _find_row(ws, "Verminderung des Bestandes an fertigen und unfertigen Erzeugnissen")
    formula = ws.cell(je_row, 3).value  # Spalte 2024 = C
    assert isinstance(formula, str)
    # Bestand-Zeile muss in der Formel addiert (+), nicht subtrahiert (−) sein
    assert f"+C{bestand_row}" in formula or formula.startswith(f"=C{bestand_row}") \
        or f"=C{bestand_row}+" in formula
    assert f"-C{bestand_row}" not in formula


def test_bwa_jue_uses_bwa_endwert_not_aggregate_sum():
    """Live-Bug 2026-05-13 (Job df31b6cd): BWA-Spalte zeigte JÜ 5,1 Mio statt
    23.585. Ursache: JÜ-Formel addierte alle Top-Level-Gruppen einer BWA inkl.
    der semantischen Aggregat-Stufen (Rohertrag = Umsatz - Material, Gesamt-
    leistung = Umsatz + Bestandsv. + Aktiv. Eigenl., Betriebsergebnis, etc.).
    Diese Aggregate sind algebraische Kombinationen ihrer Komponenten, die als
    eigene Top-Levels parallel existieren → Mehrfach-Zählung.

    Fix: Für BWA-Spalten KEINE JÜ-Formel über Top-Levels. Stattdessen direkte
    Referenz auf den BWA-internen Endwert (Vorläufiges Ergebnis o.ä.). Damit
    matcht die BWA-Spalten-JÜ genau dem in der BWA stehenden Wert.
    """
    cons = {
        "columns": [
            {"label": "BWA 2025", "kind": "bwa", "year": 2025,
             "sign_convention": "expenses_positive"},
        ],
        "groups": [
            {"name": "Umsatzerlöse", "type": "ertrag",
             "gkv_section": "umsatzerloese", "sub_group_of": None,
             "column_sums": {0: 3000000}, "accounts": []},
            {"name": "Gesamtleistung", "type": "ertrag",
             "gkv_section": "neutral", "sub_group_of": None,
             "column_sums": {0: 3000000}, "accounts": []},  # Aggregat von Umsatz
            {"name": "Material-/Wareneinkauf", "type": "aufwand",
             "gkv_section": "neutral", "sub_group_of": None,
             "column_sums": {0: 2000000}, "accounts": []},
            {"name": "Rohertrag", "type": "neutral",
             "gkv_section": "neutral", "sub_group_of": None,
             "column_sums": {0: 1000000}, "accounts": []},  # Aggregat
            {"name": "Personalkosten", "type": "aufwand",
             "gkv_section": "neutral", "sub_group_of": None,
             "column_sums": {0: 800000}, "accounts": []},
            {"name": "Betriebsergebnis", "type": "neutral",
             "gkv_section": "neutral", "sub_group_of": None,
             "column_sums": {0: 200000}, "accounts": []},  # Aggregat
            {"name": "Steuern Einkommen u. Ertrag", "type": "steuer",
             "gkv_section": "neutral", "sub_group_of": None,
             "column_sums": {0: 50000}, "accounts": []},
            {"name": "Vorläufiges Ergebnis", "type": "neutral",
             "gkv_section": "neutral", "sub_group_of": None,
             "column_sums": {0: 150000}, "accounts": []},  # echter BWA-Endwert
        ],
        "questions": [],
    }
    xlsx = build_excel(cons)
    ws = _ws(xlsx)
    je_row = _find_row(ws, "Jahresergebnis")
    endwert_row = _find_row(ws, "Vorläufiges Ergebnis")
    formula = ws.cell(je_row, 3).value
    # JÜ-Zelle: direkte Referenz auf BWA-Endwert "=Cxx", nicht "=C2+C3+..."
    expected = f"=C{endwert_row}"
    assert formula == expected, \
        f"BWA-JÜ sollte direkter Verweis auf BWA-Endwert sein " \
        f"({expected}), ist {formula!r}. " \
        f"Vermutlich Aggregat-Doppelzählung wieder eingeführt."


def test_alle_gruppen_sum_zellen_sind_formeln_kein_hardcoded_wert():
    """CLAUDE.md-Regel-Enforcement: 'Alle Excel-Zwischensummen MÜSSEN Formeln
    sein (=SUM(...) oder Kaskaden). Niemals hardcoded Werte.'

    Scannt nach build_excel jede Gruppen-Sum-Zelle (= Top-Level UND Sub-Group)
    in jeder JA-Spalte: Wert MUSS String sein und mit '=' beginnen — oder
    None (Parent-ohne-children-but-with-children-Pattern bevor Pass 2). Numeric
    Werte in Gruppen-Sum-Zellen sind verboten.

    Bilanzbericht-Format-Edge-Case: Wenn eine Gruppe accounts hat und
    pdf_sum_gj != acc_sum, muss ein Restposten-Konto ergänzt werden — die
    Sum-Zelle bleibt Formel. NICHT column_sum direkt schreiben (Live-Bug
    2026-05-13).

    AUSNAHME: BWA-Spalten (kind='bwa') dürfen für Aggregat-Gruppen ohne
    accounts und ohne children einen direkten Wert haben (BWA-Aggregate
    haben oft keine Hierarchie); aber für Gruppen MIT accounts gilt auch
    hier die Formel-Pflicht.
    """
    # Stelle Bilanzbericht-Edge-Case nach: column_sum != acc_sum
    data = {
        "columns": [
            {"label": "2024", "kind": "ja", "year": 2024,
             "sign_convention": "expenses_positive"},
        ],
        "groups": [
            {"name": "Aufwendungen", "type": "aufwand", "sub_group_of": None,
             "gkv_section": "sonst_betr_aufw",
             # column_sum WEICHT von acc_sum (1000) AB → vorher Bug: Direktwert,
             # jetzt: Restposten ergänzt, Sum-Zelle bleibt Formel
             "column_sums": {0: 5000},
             "accounts": [{"konto_nr": "6300", "bezeichnung": "Test",
                            "values": {0: 1000}, "confidence": "high"}]},
        ],
        "questions": [],
    }
    xlsx = build_excel(data)
    ws = _ws(xlsx)
    aufw_row = _find_row(ws, "Aufwendungen")
    cell_val = ws.cell(aufw_row, 3).value
    assert isinstance(cell_val, str) and cell_val.startswith("="), (
        f"REGEL VERLETZT: Gruppen-Sum-Zelle muss Formel sein, ist "
        f"{cell_val!r} (Typ {type(cell_val).__name__}). "
        f"CLAUDE.md-Regel: 'Alle Excel-Zwischensummen MÜSSEN Formeln sein'."
    )


def test_bwa_jue_leer_wenn_kein_endwert_erkennbar():
    """Wenn die BWA-Daten keinen erkennbaren Endwert-Namen ("Vorläufiges
    Ergebnis" o.ä.) haben, bleibt die JÜ-Zelle der BWA-Spalte leer — besser
    als eine falsche Aggregat-Summe."""
    cons = {
        "columns": [
            {"label": "BWA 2025", "kind": "bwa", "year": 2025,
             "sign_convention": "expenses_positive"},
        ],
        "groups": [
            {"name": "Umsatzerlöse", "type": "ertrag",
             "gkv_section": "umsatzerloese", "sub_group_of": None,
             "column_sums": {0: 3000000}, "accounts": []},
        ],
        "questions": [],
    }
    xlsx = build_excel(cons)
    ws = _ws(xlsx)
    je_row = _find_row(ws, "Jahresergebnis")
    assert ws.cell(je_row, 3).value is None, \
        "BWA-JÜ ohne erkennbaren Endwert sollte leer sein"


# ---------------------------------------------------------------------------
# EÜR (§4 Abs 3 EStG) — Karstens-Pattern
# ---------------------------------------------------------------------------

def _eur_karstens_2024():
    """Vereinfachter EÜR-Datensatz nach Karstens-Layout:
       A. Betriebseinnahmen → ertrag
       B. Betriebsausgaben → aufwand
       D.1 Hinzurechnungen → ertrag (addiert sich zum Gewinn)
       D. Kürzungen → aufwand (mindert Gewinn)

    Erwartete Excel-Formel (expenses_positive):
       Steuerlicher Gewinn = (Einnahmen + Hinzurechnungen)
                             - (Materialausgaben + Kürzungen)
       = (372474.75 + (543.10 + (-6832.00))) - (722.71 + 69483.50)
       = 366185.85 - 70206.21 = 295979.64
    """
    cols = [{"label": "2024", "kind": "ja", "year": 2024,
             "sign_convention": "expenses_positive"}]
    groups = [
        {"name": "A. 1. Einnahmen", "type": "ertrag", "sub_group_of": None,
         "gkv_section": "umsatzerloese", "column_sums": {},
         "accounts": [
             {"konto_nr": "8400", "bezeichnung": "Erlöse 19% USt",
              "values": {0: 372474.75}, "confidence": "high"},
         ]},
        {"name": "B. 1. Materialausgaben", "type": "aufwand",
         "sub_group_of": None, "gkv_section": "materialaufwand_rhb",
         "column_sums": {},
         "accounts": [
             {"konto_nr": "1600", "bezeichnung": "Verbindlichkeiten L+L",
              "values": {0: 722.71}, "confidence": "high"},
         ]},
        {"name": "D. 1. Hinzurechnungen", "type": "ertrag",
         "sub_group_of": None, "gkv_section": None, "column_sums": {},
         "accounts": [
             {"konto_nr": "4654", "bezeichnung": "Bewirtungskosten",
              "values": {0: 543.10}, "confidence": "high"},
             {"konto_nr": "4320", "bezeichnung": "Gewerbesteuer",
              "values": {0: -6832.00}, "confidence": "high"},
         ]},
        {"name": "D. Kürzungen", "type": "aufwand",
         "sub_group_of": None, "gkv_section": None, "column_sums": {},
         "accounts": [
             {"konto_nr": "9971", "bezeichnung": "IAB §7g (1) EStG",
              "values": {0: 69483.50}, "confidence": "high"},
         ]},
    ]
    return {"columns": cols, "groups": groups, "questions": [],
            "endwert_label": "Steuerlicher Gewinn nach §4 Abs. 3 EStG"}


def test_eur_endwert_label_in_excel():
    """Excel-Formel-Zeile beschriftet sich mit dem EÜR-Endwert. HGB-Default
    bleibt 'Jahresergebnis' (Backwards-Kompat)."""
    xlsx = build_excel(_eur_karstens_2024())
    ws = _ws(xlsx)
    assert _find_row(ws, "Steuerlicher Gewinn nach §4 Abs. 3 EStG") is not None
    xlsx_hgb = build_excel(_sample())
    ws_hgb = _ws(xlsx_hgb)
    assert _find_row(ws_hgb, "Jahresergebnis") is not None


def test_eur_hinzurechnungen_addiert_kuerzungen_subtrahiert():
    """Kern-Logik der EÜR: Hinzurechnungen werden im Endwert ADDIERT
    (type=ertrag), Kürzungen SUBTRAHIERT (type=aufwand). Damit kommt bei
    expenses_positive die richtige steuerliche Gewinn-Formel raus."""
    xlsx = build_excel(_eur_karstens_2024())
    ws = _ws(xlsx)
    je_row = _find_row(ws, "Steuerlicher Gewinn nach §4 Abs. 3 EStG")
    formula = ws.cell(je_row, 3).value
    einnahmen_row = _find_row(ws, "A. 1. Einnahmen")
    material_row = _find_row(ws, "B. 1. Materialausgaben")
    hinzu_row = _find_row(ws, "D. 1. Hinzurechnungen")
    kuerz_row = _find_row(ws, "D. Kürzungen")
    # Einnahmen + Hinzurechnungen sind Ertrag → in Plus-Teil
    assert f"C{einnahmen_row}+C{hinzu_row}" in formula \
        or f"C{hinzu_row}+C{einnahmen_row}" in formula
    # Material + Kürzungen sind Aufwand → in Minus-Teil (mit '-' Prefix)
    assert f"-C{material_row}" in formula
    assert f"-C{kuerz_row}" in formula


def test_eur_pdf_anker_label_dynamisch():
    """Wenn pdf_jue_per_column gesetzt ist, beschriftet die Anker-Zeile
    sich mit 'PDF-{endwert_label}' (statt hartcoded 'PDF-Jahresüberschuss')."""
    data = _eur_karstens_2024()
    # Steuerlicher Gewinn = (372474.75 - 6288.90) - (722.71 + 69483.50)
    #                     = 295979.64
    data["pdf_jue_per_column"] = {0: 295979.64}
    xlsx = build_excel(data)
    ws = _ws(xlsx)
    label_row = _find_row(ws, "PDF-Steuerlicher Gewinn nach §4 Abs. 3 EStG")
    assert label_row is not None
    assert abs(ws.cell(label_row, 3).value - 295979.64) < 0.01


def test_group_with_accounts_falls_back_to_column_sum_when_column_empty():
    """Live-Bug 2026-05-12 Edge-Case: Eine Gruppe hat accounts (z.B. "6. Zinsen
    und ähnliche Aufwendungen" mit 3 Konten aus älteren JAs), aber im jüngsten
    JA stehen keine Einzelkonten — nur pdf_sum_gj. Konsolidiert: 3 Konten mit
    Werten in 2021-2023, alle None für 2024, column_sums[2024]=4912.98.

    Vorher: has_details=True (accounts existieren) → SUM-Formel überall.
    Für 2024-Spalte: SUM(0,0,0)=0 statt 4912.98 → Diff in JÜ-Formel.

    Fix: Wenn accounts in dieser Spalte ALLE leer (kein numerischer Wert),
    column_sum-Wert direkt verwenden statt SUM(leere Zellen)=0.
    """
    cols = [
        {"label": "2023", "kind": "ja", "year": 2023, "sign_convention": "expenses_positive"},
        {"label": "2024", "kind": "ja", "year": 2024, "sign_convention": "expenses_positive"},
    ]
    groups = [
        {"name": "Umsatzerlöse", "type": "ertrag", "sub_group_of": None,
         "gkv_section": "umsatzerloese", "column_sums": {},
         "accounts": [{"konto_nr": "8400", "bezeichnung": "Erlöse",
                       "values": {0: 100000, 1: 110000}, "confidence": "high"}]},
        {"name": "6. Zinsen", "type": "aufwand", "sub_group_of": None,
         "gkv_section": "zinsaufwand",
         "column_sums": {1: 4912.98},  # nur 2024 hat pdf_sum_gj
         "accounts": [
             {"konto_nr": "7320", "bezeichnung": "Zinsen lfr.",
              "values": {0: 4511.97}, "confidence": "high"},  # nur 2023 hat Konten-Wert
         ]},
    ]
    data = {"columns": cols, "groups": groups, "questions": [],
            "endwert_label": "Jahresüberschuss"}
    xlsx = build_excel(data)
    ws = _ws(xlsx)
    zinsen_row = _find_row(ws, "6. Zinsen")
    # BEIDE Spalten: SUM-Formel (Restposten ergänzt für 2024)
    for col in (3, 4):
        v = ws.cell(zinsen_row, col).value
        assert isinstance(v, str) and v.startswith("=SUM"), \
            f"Spalte {col} sollte SUM-Formel haben, ist {v!r}"
    # Restposten-Zeile mit 2024-Wert = 4912.98 (acc_sum=0 → diff=column_sum)
    restposten_row = _find_row(ws, "  Restposten — nicht aufgeschlüsselt im PDF")
    assert restposten_row is not None, "Restposten-Konto sollte ergänzt sein"
    assert ws.cell(restposten_row, 4).value == 4912.98, \
        f"Restposten 2024 sollte 4912.98 sein, ist {ws.cell(restposten_row, 4).value!r}"


def test_rohergebnis_format_uses_column_sums_when_no_accounts():
    """Live-Bug 2026-05-12 (Job df31b6cd): DATEV-Rohergebnis-Format-JAs
    liefern fast alle GuV-Top-Levels OHNE einzelne Konten — nur pdf_sum_gj
    pro Gruppe. Beispiel:
      - "1. Rohergebnis" type=ertrag, accounts=0, pdf_sum_gj=1.190.321,13
      - "2. Personalaufwand a) Löhne" type=aufwand, accounts=0,
        pdf_sum_gj=560.694,45, sub_of="2. Personalaufwand"
      - "4. sonstige betriebliche Aufwendungen" type=aufwand, accounts>0

    Vorher: Sub-Zelle wird auf 0 gesetzt (else-Branch im Pass 1), JÜ-Formel
    überspringt Rohergebnis-Top-Level (is_bare_top_level-Heuristik) → JÜ-Excel
    extrem falsch, Build-Time-Cross-Check fail. User sieht "Verarbeitung
    fehlgeschlagen".

    Fix: Wenn accounts=0 aber column_sums[col_idx] gesetzt ist, den Wert
    in die Zelle schreiben + in die JÜ-Formel einbeziehen.
    """
    cols = [{"label": "2024", "kind": "ja", "year": 2024,
             "sign_convention": "expenses_positive"}]
    groups = [
        {"name": "1. Rohergebnis", "type": "ertrag", "sub_group_of": None,
         "gkv_section": "umsatzerloese",
         "column_sums": {0: 1190321.13}, "accounts": []},
        {"name": "2. Personalaufwand", "type": "aufwand", "sub_group_of": None,
         "gkv_section": "personalaufwand_loehne",
         "column_sums": {}, "accounts": []},
        {"name": "2. Personalaufwand a) Löhne", "type": "aufwand",
         "sub_group_of": "2. Personalaufwand",
         "gkv_section": "personalaufwand_loehne",
         "column_sums": {0: 560694.45}, "accounts": []},
        {"name": "2. Personalaufwand b) Sozial", "type": "aufwand",
         "sub_group_of": "2. Personalaufwand",
         "gkv_section": "personalaufwand_sozial",
         "column_sums": {0: 65357.11}, "accounts": []},
        {"name": "4. sonstige betriebliche Aufwendungen", "type": "aufwand",
         "sub_group_of": None, "gkv_section": "sonst_betr_aufw",
         "column_sums": {0: 214284.17},
         "accounts": [
             {"konto_nr": "6300", "bezeichnung": "div. Aufw.",
              "values": {0: 214284.17}, "confidence": "high"},
         ]},
    ]
    # Erwarteter Endwert: 1190321.13 - 560694.45 - 65357.11 - 214284.17
    #                   = 349985.40
    data = {"columns": cols, "groups": groups, "questions": [],
            "pdf_jue_per_column": {0: 349985.40},
            "endwert_label": "Jahresüberschuss"}
    # Darf nicht mit Cross-Check-ValueError fehlschlagen
    xlsx = build_excel(data)
    ws = _ws(xlsx)
    # Rohergebnis-Zelle: column_sum-Wert, nicht 0
    rohe_row = _find_row(ws, "1. Rohergebnis")
    assert ws.cell(rohe_row, 3).value == 1190321.13, \
        f"Rohergebnis-Zelle sollte 1.190.321,13 sein, ist {ws.cell(rohe_row, 3).value}"
    # 2.a Löhne-Zelle: column_sum-Wert, nicht 0
    loehne_row = _find_row(ws, "  2. Personalaufwand a) Löhne")
    assert ws.cell(loehne_row, 3).value == 560694.45, \
        f"Löhne-Zelle sollte 560.694,45 sein, ist {ws.cell(loehne_row, 3).value}"


# --- Phase 3b: manuell nachgetragene Konten ---

def test_apply_manual_accounts_schliesst_luecke_kein_restposten():
    from app.excel.builder import _apply_manual_accounts
    groups = [{"name": "8. Abschreibungen", "column_sums": {0: 1000.0},
               "accounts": [{"konto_nr": "4830", "bezeichnung": "AfA",
                             "values": {0: 900.0}}]}]
    cols = [{"label": "2024", "year": 2024}]
    g2 = _apply_manual_accounts(groups, [{"group": "8. Abschreibungen", "col_idx": 0,
                                          "bezeichnung": "AfA GWG", "betrag": 100.0}], cols)
    names = [a["bezeichnung"] for a in g2[0]["accounts"]]
    assert "AfA GWG" in names
    # acc_sum jetzt 1000 == pdf_sum → Restposten ergänzt nichts mehr
    g3 = _inject_restposten_accounts(g2)
    rest = [a for a in g3[0]["accounts"] if a.get("confidence") == "synthetic"]
    assert rest == []


def test_apply_manual_accounts_teilbetrag_restposten_schrumpft():
    from app.excel.builder import _apply_manual_accounts
    groups = [{"name": "8. Abschreibungen", "column_sums": {0: 1000.0},
               "accounts": [{"konto_nr": "4830", "bezeichnung": "AfA",
                             "values": {0: 900.0}}]}]
    cols = [{"label": "2024", "year": 2024}]
    g2 = _apply_manual_accounts(groups, [{"group": "8. Abschreibungen", "col_idx": 0,
                                          "bezeichnung": "AfA GWG", "betrag": 60.0}], cols)
    g3 = _inject_restposten_accounts(g2)
    rest = [a for a in g3[0]["accounts"] if a.get("confidence") == "synthetic"]
    assert len(rest) == 1
    assert rest[0]["values"][0] == 40.0  # 1000 - (900+60)


def test_apply_manual_accounts_ignoriert_ungueltige():
    from app.excel.builder import _apply_manual_accounts
    groups = [{"name": "8. Abschreibungen", "column_sums": {0: 1000.0},
               "accounts": [{"konto_nr": "4830", "bezeichnung": "AfA",
                             "values": {0: 900.0}}]}]
    cols = [{"label": "2024", "year": 2024}]
    g2 = _apply_manual_accounts(groups, [
        {"group": "GIBTSNICHT", "col_idx": 0, "betrag": 50.0},   # unbekannte Gruppe
        {"group": "8. Abschreibungen", "col_idx": 9, "betrag": 50.0},  # col out of range
        {"group": "8. Abschreibungen", "col_idx": 0, "betrag": None},  # kein Betrag
    ], cols)
    assert len(g2[0]["accounts"]) == 1  # nichts ergänzt


def test_apply_manual_accounts_leer_no_op():
    from app.excel.builder import _apply_manual_accounts
    groups = [{"name": "X", "accounts": []}]
    assert _apply_manual_accounts(groups, [], []) is groups


def test_build_excel_manual_account_flow_kein_restposten():
    """End-to-end: review_answers['_manual_accounts'] fließt durch build_excel,
    das manuelle Konto landet im Sheet, schließt die Lücke exakt → kein
    Restposten, Gruppen-Summe bleibt Formel."""
    consolidated = {
        "columns": [{"label": "2024", "kind": "ja", "year": 2024,
                     "sign_convention": "expenses_negative"}],
        "groups": [{"name": "8. Abschreibungen", "gkv_section": "abschreibungen",
                    "type": "aufwand", "column_sums": {0: 1000.0},
                    "accounts": [{"konto_nr": "4830", "bezeichnung": "AfA",
                                  "values": {0: 900.0}}]}],
    }
    xlsx = build_excel(consolidated, review_answers={"_manual_accounts": [
        {"group": "8. Abschreibungen", "col_idx": 0,
         "bezeichnung": "AfA GWG", "betrag": 100.0}]})
    wb = load_workbook(io.BytesIO(xlsx))
    ws = wb.active
    bez = [ws.cell(r, 2).value for r in range(1, ws.max_row + 1)]
    assert any(c and "AfA GWG" in str(c) for c in bez)  # ggf. eingerückt
    assert not any(c and "Restposten" in str(c) for c in bez)


# --- Phase 3b: Fragen-Sheet nach manueller Korrektur reconcilen ---
# reconcile rechnet das ECHTE Residuum der Ziel-Gruppe (column_sum - eigene
# Konten inkl. manuell, ohne Restposten) und keyt per eindeutigem gap_index.

def _grp(acc_900, manual=None):
    accs = [{"konto_nr": "4830", "values": {0: acc_900}}]
    if manual is not None:
        accs.append({"bezeichnung": "manuell", "values": {0: manual},
                     "confidence": "manual"})
    return [{"name": "8. Abschreibungen", "column_sums": {0: 1000.0}, "accounts": accs}]

def _gap_q():
    return [{"type": "completeness_gap", "group": "Abschreibungen", "year": 2024,
             "printed_sum": 1000.0, "acc_sum": 900.0, "diff": 100.0}]

def _man(betrag, gap_index=0):
    return {"group": "8. Abschreibungen", "col_idx": 0, "betrag": betrag,
            "gap_index": gap_index}

def test_reconcile_dropt_wenn_residuum_null():
    from app.excel.builder import _reconcile_completeness_questions
    out = _reconcile_completeness_questions(_gap_q(), _grp(900.0, manual=100.0), [_man(100.0)])
    assert out == []

def test_reconcile_teilbetrag_aktualisiert_aus_residuum():
    from app.excel.builder import _reconcile_completeness_questions
    out = _reconcile_completeness_questions(_gap_q(), _grp(900.0, manual=60.0), [_man(60.0)])
    assert len(out) == 1
    assert out[0]["diff"] == 40.0          # 1000 - (900+60)
    assert out[0]["acc_sum"] == 960.0

def test_reconcile_unangetastet_ohne_manual():
    from app.excel.builder import _reconcile_completeness_questions
    assert _reconcile_completeness_questions(_gap_q(), _grp(900.0), []) == _gap_q()

def test_build_excel_fragen_konsistent_zu_review_panel():
    """Codex P2: das Fragen-Sheet muss zur Review-Panel-Liste konsistent sein.
    Eine korrigierte Lücke verschwindet, eine unkorrigierte bleibt, andere
    Fragen (unmatched_account) überleben — kein stale completeness_gap."""
    from app.completeness import completeness_gaps
    consolidated = {
        "columns": [{"label": "2024", "kind": "ja", "doc_type": "ja", "year": 2024,
                     "sign_convention": "expenses_negative"}],
        "groups": [
            {"name": "8. Abschreibungen", "gkv_section": "abschreibungen",
             "type": "aufwand", "column_sums": {0: 1000.0},
             "accounts": [{"konto_nr": "4830", "bezeichnung": "AfA", "values": {0: 900.0}}]},
            {"name": "7. Sonstige", "gkv_section": "sonst_betr_aufw",
             "type": "aufwand", "column_sums": {0: 500.0},
             "accounts": [{"konto_nr": "6800", "bezeichnung": "Büro", "values": {0: 400.0}}]},
        ],
        "questions": [
            {"type": "unmatched_account", "konto_nr": "9", "bezeichnung": "?", "year": 2024},
            {"type": "completeness_gap", "group": "8. Abschreibungen", "year": 2024,
             "printed_sum": 1000.0, "acc_sum": 900.0, "diff": 100.0},
            {"type": "completeness_gap", "group": "7. Sonstige", "year": 2024,
             "printed_sum": 500.0, "acc_sum": 400.0, "diff": 100.0},
        ],
    }
    # canonical: zwei Lücken, in Reihenfolge → gap_index 0 = Abschreibungen
    cg = completeness_gaps(consolidated)
    assert [g["group"] for g in cg] == ["8. Abschreibungen", "7. Sonstige"]
    # User korrigiert NUR gap_index 0 (Abschreibungen) voll
    xlsx = build_excel(consolidated, review_answers={"_manual_accounts": [
        {"group": "8. Abschreibungen", "col_idx": 0, "bezeichnung": "AfA GWG",
         "betrag": 100.0, "gap_index": 0}]})
    wb = load_workbook(io.BytesIO(xlsx))
    fragen = "\n".join(str(c.value) for row in wb["Fragen"].iter_rows()
                       for c in row if c.value)
    assert "8. Abschreibungen" not in fragen   # korrigiert → weg
    assert "7. Sonstige" in fragen             # unkorrigiert → bleibt
    assert "Konto 9" in fragen or "'?'" in fragen  # unmatched_account überlebt

def test_reconcile_keyed_per_index_nicht_gruppe():
    """Codex P3: zwei Lücken mit gleichem Gruppe+Jahr — eine Korrektur (gap_index
    0) darf nur DIESE Warnung entfernen, nicht beide."""
    from app.excel.builder import _reconcile_completeness_questions
    q = [{"type": "completeness_gap", "group": "A", "year": 2024,
          "printed_sum": 100.0, "acc_sum": 0.0, "diff": 100.0},
         {"type": "completeness_gap", "group": "A", "year": 2024,
          "printed_sum": 100.0, "acc_sum": 0.0, "diff": 100.0}]
    groups = [{"name": "A", "column_sums": {0: 100.0},
               "accounts": [{"values": {0: 100.0}, "confidence": "manual"}]}]
    out = _reconcile_completeness_questions(q, groups, [{"group": "A", "col_idx": 0,
                                                         "betrag": 100.0, "gap_index": 0}])
    assert len(out) == 1  # nur die korrigierte (index 0) raus

def test_build_excel_manual_account_auf_parent_wird_ignoriert():
    """Codex P1: ein manuelles Konto auf einem Parent (mit Sub-Gruppen) würde
    dessen Sum-Rendering kippen + Kinder doppelt zählen. Parent-Targets müssen
    verworfen werden — das Konto darf NICHT auftauchen, Parent bleibt Kaskade."""
    consolidated = {
        "columns": [{"label": "2024", "kind": "ja", "year": 2024,
                     "sign_convention": "expenses_negative"}],
        "groups": [
            {"name": "4. Materialaufwand", "gkv_section": "materialaufwand_rhb",
             "type": "aufwand", "sub_group_of": None,
             "column_sums": {0: 1000.0}, "accounts": []},
            {"name": "4. a) RHB", "type": "aufwand", "sub_group_of": "4. Materialaufwand",
             "column_sums": {0: 1000.0},
             "accounts": [{"konto_nr": "5100", "bezeichnung": "RHB", "values": {0: 900.0}}]},
        ],
    }
    xlsx = build_excel(consolidated, review_answers={"_manual_accounts": [
        {"group": "4. Materialaufwand", "col_idx": 0, "bezeichnung": "PARENT-KONTO",
         "betrag": 100.0, "gap_group_orig": "4. Materialaufwand", "gap_year_orig": "2024"}]})
    ws = load_workbook(io.BytesIO(xlsx)).active
    bez = [ws.cell(r, 2).value for r in range(1, ws.max_row + 1)]
    assert not any(c and "PARENT-KONTO" in str(c) for c in bez), \
        "manuelles Konto darf nicht in eine Parent-Gruppe injiziert werden"
    # Parent-Sum referenziert die Kinder-Sumrow (Kaskade), kein =SUM eigener Konten
    parent_row = _find_row(ws, "4. Materialaufwand")
    child_row = _find_row(ws, "  4. a) RHB")
    assert parent_row and child_row
    pf = ws.cell(parent_row, 3).value
    assert isinstance(pf, str) and f"C{child_row}" in pf


def test_match_group_name_mehrdeutig_gibt_none():
    """Code-Review R5: bidirektionaler Substring-Match darf bei MEHREREN Treffern
    nicht raten — sonst bindet die Lücke an die falsche Gruppe (recompute → ggf.
    stilles Droppen). >1 Treffer → None (User wählt selbst)."""
    from app.completeness import _match_group_name
    assert _match_group_name("Abschreibungen",
                             ["7. Abschreibungen Sachanlagen", "7a. Abschreibungen GWG"]) is None
    assert _match_group_name("Abschreibungen",
                             ["8. Abschreibungen", "1. Umsatz"]) == "8. Abschreibungen"
    # exakter Match gewinnt auch bei zusätzlichem Substring-Kandidaten
    assert _match_group_name("8. X", ["8. X", "8. X Sub"]) == "8. X"


def test_reconcile_updated_gap_konsistent_zur_gruppe():
    """Code-Review R5: bei Teilkorrektur kommen printed_sum/acc_sum/diff aus der
    KONSOLIDIERTEN Gruppe (nicht aus stale gap.printed_sum 'or 0.0'). Sonst
    negativer/inkonsistenter acc_sum im Fragen-Sheet."""
    from app.excel.builder import _reconcile_completeness_questions
    questions = [{"type": "completeness_gap", "group": "Abschreibungen", "year": 2024,
                  "printed_sum": None, "acc_sum": None, "diff": 100.0}]
    out = _reconcile_completeness_questions(questions, _grp(900.0, manual=60.0), [_man(60.0)])
    assert out[0]["printed_sum"] == 1000.0
    assert out[0]["acc_sum"] == 960.0
    assert out[0]["diff"] == 40.0
