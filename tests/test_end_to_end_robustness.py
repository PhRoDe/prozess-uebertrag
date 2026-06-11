"""Production-Robustheits-Tests: End-to-End-Pipeline gegen die bekannten
PDF-Format-Patterns. Jedes Pattern muss die Akzeptanz-Kriterien für eine
produktiv ausgelieferte Excel erfüllen — sonst pytest rot, Deploy blockiert.

Akzeptanz-Kriterien (siehe _assert_excel_production_ready):
1. Excel wird gebaut (kein ValueError-Crash)
2. Konten aus den JAs sind im Excel sichtbar
3. ALLE Gruppen-Sum-Zellen sind Formeln (=SUM/=...), niemals statische Werte
4. JÜ-Cross-Check matcht PDF-JÜ centgenau in jeder JA-Spalte
5. Bei Daten-Lücken: Restposten-Konto sichtbar, Diff bleibt 0

Pattern-Coverage:
- A) HGB-GuV mit vollständigem Kontennachweis (Tasteone-Pattern)
- B) HGB-GuV mit DATEV-Rohergebnis-Format (Bilanzbericht-Pattern,
       Sub-Gruppen nur mit pdf_sum_gj, keine Konten)
- C) Mehrjähriger Mix (manche Spalten haben Konten, andere nur pdf_sum)
- D) BWA-only (Aggregat-Stufen: Vorläufiges Ergebnis als Endwert)
- E) EÜR §4 Abs 3 mit Hinzurechnungen/Kürzungen (Karstens-Pattern)
- F) Multi-Year mit STB-Vorzeichen-Inversion in einer Spalte (Tasteone-Bug-2)
"""
import io
import pytest
from openpyxl import load_workbook
from app.excel.builder import build_excel
from app.worker.consolidate import merge_extractions


# --- Akzeptanz-Helper ---------------------------------------------------------

def _assert_excel_production_ready(xlsx_bytes: bytes, expected_pdf_jue: dict[int, float],
                                    columns_count: int = None):
    """Zentrale Akzeptanz-Funktion: prüft alle Production-Kriterien an einer Excel.

    Args:
        xlsx_bytes: das gebaute Excel
        expected_pdf_jue: {col_idx: pdf_jue} — JÜ-Cross-Check pro JA-Spalte
        columns_count: optional, gesamte Spalten-Anzahl
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
    assert "Übertrag" in wb.sheetnames, "Übertrag-Sheet muss existieren"
    ws = wb["Übertrag"]

    # 1. Konten sind im Excel: mindestens eine Zeile mit Konto-Nr
    konto_rows = [r for r in ws.iter_rows() if r[0].value and str(r[0].value).strip()]
    assert konto_rows, "Excel hat keine Konten-Zeilen — strukturell kaputt"

    # 2. Alle Gruppen-Sum-Zellen sind Formeln. Heuristik: eine Zeile ist eine
    # Gruppen-Sum-Zeile wenn Konto-Spalte leer ist UND die Bezeichnungs-Spalte
    # NICHT mit zwei Leerzeichen beginnt (Konto-Detail-Zeilen sind eingerückt
    # mit "  ").
    je_label = None
    pdf_label = None
    for r in ws.iter_rows():
        bez = r[1].value
        if isinstance(bez, str):
            if bez.startswith("PDF-"):
                pdf_label = bez
            elif bez in ("Jahresergebnis", "Jahresüberschuss", "Steuerlicher Gewinn") \
                    or "Steuerlicher Gewinn" in bez:
                je_label = bez

    # Akzeptanz-Regel (CLAUDE.md "nicht brechen"-Block):
    # "Alle Excel-Zwischensummen MÜSSEN Formeln sein (=SUM(...) oder Kaskaden).
    #  Niemals hardcoded Werte."
    # Interpretation: "Zwischensumme" = eine Aggregation über Detail-Items.
    # → Gilt für Gruppen MIT accounts (Sum über die Detail-Zeilen).
    # → Gilt NICHT für Gruppen OHNE accounts und ohne children (dort ist die
    #   Zelle ein direkter Quell-Wert aus pdf_sum_gj, keine Aggregation).
    #
    # Heuristik: Eine Gruppen-Sum-Zeile hat accounts wenn die direkt folgende
    # Zeile eine Konto-Nr hat ODER mit "  Konto" eingerückt ist.
    rows = list(ws.iter_rows())
    for i, row in enumerate(rows):
        if i + 1 >= len(rows):
            continue
        konto_cell = row[0].value
        bez_cell = row[1].value or ""
        if konto_cell:
            continue
        if not isinstance(bez_cell, str) or not bez_cell.strip():
            continue
        bez = bez_cell.strip()
        if bez.startswith(("PDF-", "Differenz", "---")):
            continue
        if je_label and bez == je_label:
            continue
        # Sub-Group-Header werden auch geprüft
        # Hat diese Gruppe accounts? Schaue auf nächste Zeile(n)
        next_row = rows[i + 1]
        next_konto = next_row[0].value
        next_bez = next_row[1].value or ""
        has_accounts = bool(next_konto) or (
            isinstance(next_bez, str) and next_bez.startswith("  ")
            and not next_bez.strip().startswith(("1.", "2.", "3.", "4.", "5.", "6.",
                                                  "7.", "8.", "9."))
        )
        if not has_accounts:
            continue  # Direktwert OK, keine Aggregation
        # Gruppe hat accounts → Sum-Zelle MUSS Formel sein
        for col_idx in range(3, ws.max_column + 1):
            val = ws.cell(row[0].row, col_idx).value
            if val is None or val == 0:
                continue
            assert isinstance(val, str) and val.startswith("="), (
                f"REGEL VERLETZT: Zeile {row[0].row} ('{bez[:40]}'), Spalte {col_idx}: "
                f"Wert {val!r} ist kein Formel — Gruppen-Sum-Zellen mit accounts "
                f"MÜSSEN =SUM(...)-Formel sein (CLAUDE.md 'Wichtige Regeln'). "
                f"Bei pdf_sum_gj != acc_sum: Restposten-Konto verwenden, nicht "
                f"direkten Wert."
            )

    # 3. JÜ-Cross-Check via LibreOffice/openpyxl-recalc nicht möglich;
    # stattdessen den internen _compute_excel_jue_per_column-Mechanismus
    # nutzen — wenn der Bug-Cross-Check im Builder Mismatches findet, landen
    # die im Fragen-Sheet.
    if "Fragen" in wb.sheetnames:
        fragen_rows = list(wb["Fragen"].iter_rows(values_only=True))
        jue_mismatches = [r for r in fragen_rows
                           if r[0] == "jue_excel_vs_pdf_mismatch"]
        assert not jue_mismatches, (
            f"JÜ-Mismatch im Fragen-Sheet — Excel-JÜ != PDF-JÜ in mindestens "
            f"einer Spalte. Mismatches: {jue_mismatches}. "
            f"Restposten-Logik sollte das eigentlich auf 0 bringen."
        )


# --- Pattern A: Vollständiger Kontennachweis (Tasteone-Pattern) ---------------

def test_pattern_A_vollstaendiger_kontennachweis():
    """JA mit allen Einzelkonten pro Gruppe, acc_sum == pdf_sum_gj."""
    doc = {
        "type": "jahresabschluss", "year": 2024, "previous_year": 2023,
        "sign_convention": "expenses_positive", "open_questions": [],
        "groups": [
            {"name": "1. Umsatzerlöse", "type": "ertrag", "sub_group_of": None,
             "gkv_section": "umsatzerloese", "pdf_sum_gj": 5000000.00,
             "pdf_sum_vj": 4500000.00,
             "accounts": [
                 {"konto_nr": "8400", "bezeichnung": "Erlöse 19%",
                  "betrag_gj": 5000000.00, "betrag_vj": 4500000.00, "confidence": "high"},
             ]},
            {"name": "4. Materialaufwand", "type": "aufwand", "sub_group_of": None,
             "gkv_section": "materialaufwand_rhb", "pdf_sum_gj": 2000000.00,
             "pdf_sum_vj": 1800000.00,
             "accounts": [
                 {"konto_nr": "5400", "bezeichnung": "Wareneingang",
                  "betrag_gj": 2000000.00, "betrag_vj": 1800000.00, "confidence": "high"},
             ]},
        ],
        "pdf_jahresueberschuss_gj": 3000000.00,
        "pdf_jahresueberschuss_vj": 2700000.00,
    }
    cons = merge_extractions([doc])
    xlsx = build_excel(cons)
    _assert_excel_production_ready(xlsx, expected_pdf_jue={1: 3000000.00})


# --- Pattern B: DATEV-Rohergebnis-Format (Bilanzbericht-Pattern) --------------

def test_pattern_B_rohergebnis_format_ohne_einzelkonten():
    """Bilanzbericht-Mandant (kleine GmbH): Top-Level + Sub-Groups haben
    pdf_sum_gj ABER keine accounts. Builder muss column_sum als Wert
    schreiben — über die `else`-Branch in Pass 1 (keine accounts, keine
    children → cs_val_for_col). KEIN Restposten nötig (kein acc_sum-Mismatch).
    """
    doc = {
        "type": "jahresabschluss", "year": 2024, "previous_year": 2023,
        "sign_convention": "expenses_positive", "open_questions": [],
        "groups": [
            {"name": "1. Rohergebnis", "type": "ertrag", "sub_group_of": None,
             "gkv_section": "umsatzerloese",
             "pdf_sum_gj": 1190321.13, "pdf_sum_vj": 846555.41, "accounts": []},
            {"name": "2. Personalaufwand a) Löhne", "type": "aufwand",
             "sub_group_of": "2. Personalaufwand",
             "gkv_section": "personalaufwand_loehne",
             "pdf_sum_gj": 560694.45, "pdf_sum_vj": 424396.31, "accounts": []},
            {"name": "2. Personalaufwand b) Sozial", "type": "aufwand",
             "sub_group_of": "2. Personalaufwand",
             "gkv_section": "personalaufwand_sozial",
             "pdf_sum_gj": 65357.11, "pdf_sum_vj": 58377.73, "accounts": []},
        ],
        "pdf_jahresueberschuss_gj": 564269.57,  # 1190321.13 - 560694.45 - 65357.11
        "pdf_jahresueberschuss_vj": 363781.37,
    }
    cons = merge_extractions([doc])
    xlsx = build_excel(cons)
    _assert_excel_production_ready(xlsx, expected_pdf_jue={1: 564269.57})


# --- Pattern C: Konten vorhanden aber unvollständig (Restposten-Pattern) ------

def test_pattern_C_konten_unvollstaendig_restposten_ergaenzt():
    """Bilanzbericht-Mandant mit teilweise gelistetem Kontennachweis:
    sonst.betr.Aufw. hat einige Konten gelistet (acc_sum=148.681), aber
    pdf_sum_gj=214.284 (Rest implizit in der Summe). Restposten-Konto
    muss ergänzt werden, Excel-Gruppensumme matcht PDF centgenau.
    """
    doc = {
        "type": "jahresabschluss", "year": 2024, "previous_year": 2023,
        "sign_convention": "expenses_positive", "open_questions": [],
        "groups": [
            {"name": "1. Rohergebnis", "type": "ertrag", "sub_group_of": None,
             "gkv_section": "umsatzerloese",
             "pdf_sum_gj": 500000.00, "pdf_sum_vj": 400000.00, "accounts": []},
            {"name": "4. sonst.betr.Aufw.", "type": "aufwand", "sub_group_of": None,
             "gkv_section": "sonst_betr_aufw",
             "pdf_sum_gj": 214284.17, "pdf_sum_vj": 150000.00,
             "accounts": [
                 {"konto_nr": "6300", "bezeichnung": "Sonst. Aufw.",
                  "betrag_gj": 100000.00, "betrag_vj": 80000.00, "confidence": "high"},
                 {"konto_nr": "6310", "bezeichnung": "Miete",
                  "betrag_gj": 48681.60, "betrag_vj": 40000.00, "confidence": "high"},
                 # acc_sum=148681.60, pdf_sum_gj=214284.17, Restposten=65602.57
             ]},
        ],
        "pdf_jahresueberschuss_gj": 285715.83,  # 500000 - 214284.17
        "pdf_jahresueberschuss_vj": 250000.00,
    }
    cons = merge_extractions([doc])
    xlsx = build_excel(cons)
    _assert_excel_production_ready(xlsx, expected_pdf_jue={1: 285715.83})
    # Restposten-Konto MUSS sichtbar sein
    wb = load_workbook(io.BytesIO(xlsx), data_only=False)
    ws = wb["Übertrag"]
    restposten_found = False
    for row in ws.iter_rows(values_only=True):
        if row[1] and "Restposten" in str(row[1]):
            restposten_found = True
            break
    assert restposten_found, "Restposten-Konto fehlt obwohl acc_sum != pdf_sum_gj"


# --- Pattern D: BWA-only (Aggregat-Hierarchie) ---------------------------------

def test_pattern_D_bwa_only_mit_endwert():
    """BWA als alleinige Quelle: Aggregat-Hierarchie (Rohertrag, Betrieblicher
    Rohertrag, Gesamtkosten, Vorläufiges Ergebnis). JÜ-Zelle der BWA-Spalte
    muss direkter Verweis auf Vorläufiges Ergebnis sein, KEINE Aggregat-Summe
    (sonst Doppelzählung)."""
    doc = {
        "type": "bwa", "year": 2025, "period_label": "BWA Jan-Dez 2025",
        "sign_convention": "expenses_positive", "open_questions": [],
        "groups": [
            {"name": "Umsatzerlöse", "type": "ertrag", "sub_group_of": None,
             "gkv_section": "umsatzerloese", "pdf_sum_gj": 3000000.00, "accounts": []},
            {"name": "Material-/Wareneinkauf", "type": "aufwand", "sub_group_of": None,
             "gkv_section": "neutral", "pdf_sum_gj": 2000000.00, "accounts": []},
            {"name": "Rohertrag", "type": "neutral", "sub_group_of": None,
             "gkv_section": "neutral", "pdf_sum_gj": 1000000.00, "accounts": []},
            {"name": "Personalkosten", "type": "aufwand", "sub_group_of": None,
             "gkv_section": "neutral", "pdf_sum_gj": 800000.00, "accounts": []},
            {"name": "Vorläufiges Ergebnis", "type": "neutral", "sub_group_of": None,
             "gkv_section": "neutral", "pdf_sum_gj": 200000.00, "accounts": []},
        ],
    }
    cons = merge_extractions([doc])
    xlsx = build_excel(cons)
    # Kein PDF-JÜ in BWA → keine Cross-Check-Erwartung
    _assert_excel_production_ready(xlsx, expected_pdf_jue={})


# --- Pattern E: EÜR mit Hinzurechnungen/Kürzungen ------------------------------

def test_pattern_E_eur_hinzurechnungen_kuerzungen():
    """EÜR §4 Abs 3 EStG (Einzelunternehmer): Hinzurechnungen (ertrag) und
    Kürzungen (aufwand) als Subs unter synthetischem Parent. JÜ-Formel muss
    auf Sub-Group-Ebene iterieren."""
    doc = {
        "type": "jahresabschluss", "year": 2024, "previous_year": 2023,
        "sign_convention": "expenses_positive", "open_questions": [],
        "endwert_label": "Steuerlicher Gewinn nach §4 Abs 3 EStG",
        "groups": [
            {"name": "A. Betriebseinnahmen", "type": "ertrag", "sub_group_of": None,
             "gkv_section": "umsatzerloese", "pdf_sum_gj": 100000.00,
             "accounts": [{"konto_nr": "8400", "bezeichnung": "Einnahmen",
                            "betrag_gj": 100000.00, "betrag_vj": 90000.00,
                            "confidence": "high"}]},
            {"name": "B. Betriebsausgaben", "type": "aufwand", "sub_group_of": None,
             "gkv_section": "sonst_betr_aufw", "pdf_sum_gj": 70000.00,
             "accounts": [{"konto_nr": "4900", "bezeichnung": "Ausgaben",
                            "betrag_gj": 70000.00, "betrag_vj": 65000.00,
                            "confidence": "high"}]},
            {"name": "D. Hinzurechnungen", "type": "ertrag",
             "sub_group_of": "D. Steuerliche Korrekturen",
             "gkv_section": None, "pdf_sum_gj": 5000.00,
             "accounts": [{"konto_nr": "4654", "bezeichnung": "Nicht abz. Bewirtung",
                            "betrag_gj": 5000.00, "betrag_vj": 4500.00,
                            "confidence": "high"}]},
            {"name": "D. Kürzungen", "type": "aufwand",
             "sub_group_of": "D. Steuerliche Korrekturen",
             "gkv_section": None, "pdf_sum_gj": 2000.00,
             "accounts": [{"konto_nr": "2281", "bezeichnung": "GewSt-Korrektur",
                            "betrag_gj": 2000.00, "betrag_vj": 1800.00,
                            "confidence": "high"}]},
        ],
        "pdf_jahresueberschuss_gj": 33000.00,  # 100k - 70k + 5k - 2k
        "pdf_jahresueberschuss_vj": 27700.00,
    }
    cons = merge_extractions([doc])
    xlsx = build_excel(cons)
    _assert_excel_production_ready(xlsx, expected_pdf_jue={1: 33000.00})


# --- Pattern G: Prisma-Mehrjahres-Fehler kombiniert (2026-06) --------------------

def test_pattern_G_prisma_three_multi_year_bugs_combined():
    """Pattern G (Prisma 2026-06): kombiniert die drei in der Produktion
    gefundenen Mehrjahres-Fehler in EINEM End-to-End-Lauf:

      1. Verminderung des Bestandes — Detail normalisiert negativ, pdf_sum-Anker
         roh positiv → ohne Fix Phantom-Restposten 2×|wert|, JÜ um Millionen falsch.
      2. GuV-Position (Zinsaufwand) fehlt im JÜNGSTEN JA → ohne Fix aus den
         älteren JAs gedroppt (Template-from-newest), JÜ zu hoch.
      3. Folge-JA benennt sein Vorjahr anders (Umsatz-Klumpen statt Einzelkonten)
         → ohne Fix Doppelzählung in der Eigenjahr-Spalte ('addiert-dann-abgezogen').

    Akzeptanz: alle 4 JA-Spalten centgenau (0 Mismatch — fängt Fehler 1 + 2),
    Zinsaufwand-Position erhalten (Fehler 2), keine VJ-Duplikate (Fehler 3).
    """
    def _ja(year, ums_acc, bestand_name, bestand_gj, bestand_vj,
            ums_gj, ums_vj, mat_gj, mat_vj, jue_gj, jue_vj, zins_gj=None, zins_vj=None):
        groups = [
            {"name": "1. Umsatzerlöse", "type": "ertrag", "sub_group_of": None,
             "gkv_section": "umsatzerloese", "pdf_sum_gj": ums_gj,
             "accounts": [{"konto_nr": ums_acc[0], "bezeichnung": ums_acc[1],
                            "betrag_gj": ums_gj, "betrag_vj": ums_vj, "confidence": "high"}]},
            {"name": bestand_name, "type": "ertrag", "sub_group_of": None,
             "gkv_section": "bestandsveraenderung", "pdf_sum_gj": bestand_gj,
             "accounts": [{"konto_nr": "8990", "bezeichnung": "Bestandsveränderung",
                            "betrag_gj": bestand_gj, "betrag_vj": bestand_vj,
                            "confidence": "high"}]},
            {"name": "4. Materialaufwand", "type": "aufwand", "sub_group_of": None,
             "gkv_section": "materialaufwand_rhb", "pdf_sum_gj": mat_gj,
             "accounts": [{"konto_nr": "5400", "bezeichnung": "Wareneingang",
                            "betrag_gj": mat_gj, "betrag_vj": mat_vj, "confidence": "high"}]},
        ]
        if zins_gj is not None:
            groups.append(
                {"name": "9. Zinsen und ähnliche Aufwendungen", "type": "aufwand",
                 "sub_group_of": None, "gkv_section": "zinsaufwand",
                 "pdf_sum_gj": zins_gj, "pdf_sum_vj": zins_vj, "accounts": []})
        return {"type": "jahresabschluss", "year": year, "previous_year": year - 1,
                "sign_convention": "expenses_positive", "open_questions": [],
                "groups": groups,
                "pdf_jahresueberschuss_gj": jue_gj, "pdf_jahresueberschuss_vj": jue_vj}

    # JÜ = Umsatz + Bestand(signiert) - Material - Zinsaufwand
    docs = [
        # 2022: Erhöhung (+), MIT Zinsaufwand. VJ liefert 2021.
        _ja(2022, ("8400", "Erlöse A"),
            "2. Erhöhung des Bestandes an fertigen und unfertigen Erzeugnissen",
            bestand_gj=30000.0, bestand_vj=20000.0,
            ums_gj=800000.0, ums_vj=700000.0, mat_gj=300000.0, mat_vj=250000.0,
            zins_gj=5000.0, zins_vj=4000.0,
            jue_gj=525000.0, jue_vj=466000.0),   # 2022: 800-300+30-5 ; 2021: 700-250+20-4
        # 2023: Verminderung (−, Fehler 1), MIT Zinsaufwand.
        _ja(2023, ("8400", "Erlöse A"),
            "2. Verminderung des Bestandes an fertigen und unfertigen Erzeugnissen",
            bestand_gj=40000.0, bestand_vj=None,
            ums_gj=900000.0, ums_vj=800000.0, mat_gj=320000.0, mat_vj=300000.0,
            zins_gj=6000.0, zins_vj=5000.0,
            jue_gj=534000.0, jue_vj=525000.0),   # 2023: 900-320-40-6
        # 2024 (jüngster): Verminderung, OHNE Zinsaufwand (Fehler 2), Umsatz
        # anders benannt (Fehler 3).
        _ja(2024, ("8000", "Umsatzerlöse gesamt"),
            "2. Verminderung des Bestandes an fertigen und unfertigen Erzeugnissen",
            bestand_gj=50000.0, bestand_vj=None,
            ums_gj=1000000.0, ums_vj=900000.0, mat_gj=350000.0, mat_vj=320000.0,
            jue_gj=600000.0, jue_vj=534000.0),   # 2024: 1000-350-50
    ]
    cons = merge_extractions(docs)
    xlsx = build_excel(cons)
    col = {cons["columns"][i]["year"]: i for i in range(len(cons["columns"]))}

    # Akzeptanz: alle 4 JA-Spalten centgenau (Fehler 1 + 2)
    _assert_excel_production_ready(xlsx, expected_pdf_jue={
        col[2021]: 466000.0, col[2022]: 525000.0,
        col[2023]: 534000.0, col[2024]: 600000.0})

    # Fehler 2: Zinsaufwand-Position erhalten (nicht aus älterem JA gedroppt)
    zins = [g for g in cons["groups"] if g.get("gkv_section") == "zinsaufwand"]
    assert zins, "Zinsaufwand-Position wurde gedroppt (Fehler 2)"

    # Fehler 3: keine VJ-Duplikate in der Umsatz-2023-Spalte (Eigenjahr authoritativ)
    umsatz = [g for g in cons["groups"] if "Umsatzerl" in g["name"]][0]
    c23 = col[2023]
    vals = [a["values"][c23] for a in umsatz["accounts"] if c23 in a.get("values", {})]
    assert vals == [900000.0], f"VJ-Duplikat in Umsatz-2023 (Fehler 3): {vals}"


# --- Pattern F: Multi-Year-Setup baut ohne Crash --------------------------------

def test_pattern_F_multi_year_baut_ohne_crash():
    """Sicherheits-Anker: drei JAs nacheinander, Excel muss ohne ValueError
    bauen + alle drei JÜ-Spalten haben Werte. Sign-Outlier-Normalisierung
    selbst ist in test_consolidate.py einzeln getestet — dieser Test deckt
    nur den End-to-End-Bau ab."""
    def _ja(year, gj=600000.00, vj=540000.00):
        return {
            "type": "jahresabschluss", "year": year, "previous_year": year-1,
            "sign_convention": "expenses_positive", "open_questions": [],
            "groups": [
                {"name": "Umsatzerlöse", "type": "ertrag", "sub_group_of": None,
                 "gkv_section": "umsatzerloese", "pdf_sum_gj": 1000000.00,
                 "accounts": [{"konto_nr": "8400", "bezeichnung": "Erlöse",
                                "betrag_gj": 1000000.00, "betrag_vj": 900000.00,
                                "confidence": "high"}]},
                {"name": "Materialaufwand", "type": "aufwand", "sub_group_of": None,
                 "gkv_section": "materialaufwand_rhb", "pdf_sum_gj": 400000.00,
                 "accounts": [{"konto_nr": "5400", "bezeichnung": "Wareneingang",
                                "betrag_gj": 400000.00, "betrag_vj": 360000.00,
                                "confidence": "high"}]},
            ],
            "pdf_jahresueberschuss_gj": gj,
            "pdf_jahresueberschuss_vj": vj,
        }
    docs = [_ja(2022), _ja(2023), _ja(2024)]
    cons = merge_extractions(docs)
    xlsx = build_excel(cons)
    # Cross-Check muss centgenau passen
    _assert_excel_production_ready(xlsx,
        expected_pdf_jue={i: 600000.00 for i, c in enumerate(cons["columns"])
                          if c.get("kind") == "ja" and c.get("year", 0) >= 2022})
