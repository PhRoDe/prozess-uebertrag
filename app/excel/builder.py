"""Excel-Builder — erzeugt die Übertrag-Excel basierend auf der
dynamischen Gruppenstruktur aus der Konsolidierung.

Layout:
- Header: Konto | Bezeichnung | Jahr-1 | Jahr-2 | BWA xyz | ...
- Pro Gruppe:
    * Summen-Zeile (fett) OBEN: Name + SUM-Formel über Detail-Zeilen (für JAs)
      bzw. direkter Wert (für BWAs)
    * Details: alle Konten darunter
- Nach allen Gruppen: `Jahresergebnis` als Formel
  (respektiert sign_convention + Gruppen-Typen)
- Zweites Sheet `Fragen`: Data-Quality-Issues
"""
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

EUR_FORMAT = '#,##0.00;-#,##0.00'
YELLOW = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
BOLD = Font(bold=True)
BOLD_LIGHT = Font(bold=True, color="555555")

# Bilanzgewinn-Sektionen: NICHT Teil der GuV / des JÜ. Werden in einem
# eigenen Block am Ende der Excel gerendert.
BILANZGEWINN_SECTIONS = {"gewinnvortrag", "ausschuettung", "bilanzgewinn"}

# Mapping von gkv_section auf JÜ-Rolle. Authoritativ ueber `type`, weil
# Claude beim type-Feld haeufiger driftet als bei der GKV-Klassifikation.
SECTION_ROLE: dict[str, str] = {
    "umsatzerloese": "ertrag",
    "bestandsveraenderung": "ertrag",  # Werte sind im consolidated normalisiert (+ Erhöhung / − Verminderung)
    "aktivierte_eigenleistungen": "ertrag",
    "sonst_betr_ertraege": "ertrag",
    "materialaufwand_rhb": "aufwand",
    "materialaufwand_bez_leistungen": "aufwand",
    "personalaufwand_loehne": "aufwand",
    "personalaufwand_sozial": "aufwand",
    "abschreibungen": "aufwand",
    "sonst_betr_aufw": "aufwand",
    "ertraege_wertpapiere": "ertrag",
    "ertraege_beteiligungen": "ertrag",
    "sonstige_zins_ertraege": "ertrag",
    "zinsaufwand": "aufwand",
    "ee_steuern": "steuer",
    "sonst_steuern": "steuer",
}


def _coerce_int_keys(groups: list[dict]) -> list[dict]:
    """JSON roundtrip via Postgres turns int dict keys into strings. Restore
    int keys for `values` and `column_sums` so the builder can index by col_idx."""
    out = []
    for g in groups:
        g2 = dict(g)
        g2["column_sums"] = {int(k): v for k, v in (g.get("column_sums") or {}).items()}
        new_accounts = []
        for acc in g.get("accounts", []) or []:
            a2 = dict(acc)
            a2["values"] = {int(k): v for k, v in (acc.get("values") or {}).items()}
            new_accounts.append(a2)
        g2["accounts"] = new_accounts
        out.append(g2)
    return out


def _resolve_role(g: dict) -> str:
    """Bestimmt fuer eine Gruppe die JUE-Rolle (ertrag/aufwand/steuer/neutral).
    gkv_section ist authoritativ, falls bekannt; sonst fallback auf type."""
    section = g.get("gkv_section")
    if section in SECTION_ROLE:
        return SECTION_ROLE[section]
    return g.get("type", "neutral")


def _endwert_groups(groups: list[dict]) -> list[dict]:
    """Liefert die Gruppen die in die Endwert-Formel einfliessen, jeweils mit
    ihrem eigenen type/section-Vorzeichen.

    Logik:
    - Top-Level mit eigenen accounts → Top-Level verwenden (Subs werden in
      Pass 2 aufaddiert; bei HGB-Pattern haben die Subs den gleichen type
      wie der Parent, also algebraisch aequivalent).
    - Top-Level ohne accounts, aber mit Sub-Gruppen → Sub-Gruppen einzeln.
      Wichtig fuer EÜR-Block "D. STEUERLICHE KORREKTUREN": Hinzurechnungen
      (type=ertrag) und Kürzungen/IAB-Bildung (type=aufwand) bekommen
      gegenteilige Vorzeichen — der Parent kann sie nicht repraesentieren.
    - Top-Level ohne accounts und ohne Subs → bleibt drin (BWA-Aggregat,
      direkt aus column_sums).
    """
    result = []
    for g in groups:
        if g.get("sub_group_of") is not None:
            continue
        if g.get("accounts"):
            result.append(g)
        else:
            subs = [s for s in groups if s.get("sub_group_of") == g["name"]]
            if subs:
                result.extend(subs)
            else:
                result.append(g)
    return result


def build_excel(consolidated: dict, review_answers: dict | None = None) -> bytes:
    """Build the final xlsx bytes from the consolidated structure.

    review_answers: optional mapping {konto_nr -> target_group_name} for accounts
                    that landed in open_questions.
    """
    review_answers = review_answers or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Übertrag"

    columns = consolidated.get("columns", [])
    groups = _coerce_int_keys(consolidated.get("groups", []))
    pdf_jue_per_column = {int(k): v for k, v in
                           (consolidated.get("pdf_jue_per_column") or {}).items()}
    questions = list(consolidated.get("questions") or [])
    # Endwert-Label aus der Extraktion (z.B. "Steuerlicher Gewinn nach §4 Abs 3
    # EStG" bei EÜR). Wenn gesetzt, nutzen wir es fuer beide Endwert-Zeilen
    # (Excel-Formel + PDF-Anker). Bei HGB-GuV ohne explicit Label bleiben die
    # historischen Defaults "Jahresergebnis" (Formel) und "PDF-Jahresüberschuss"
    # (Anker) — neutral genug fuer Ueberschuss UND Fehlbetrag.
    endwert_label_explicit = consolidated.get("endwert_label")
    formula_label = endwert_label_explicit or "Jahresergebnis"
    pdf_anker_label = endwert_label_explicit or "Jahresüberschuss"

    # Sanity: if any accounts exist at all, at least one must carry a value.
    # An all-empty account set means a structural bug (typically a JSON-roundtrip
    # key-type drift) — fail loud rather than ship a blank workbook.
    accounts_total = sum(len(g.get("accounts", []) or []) for g in groups)
    accounts_with_values = sum(
        1 for g in groups for acc in (g.get("accounts") or [])
        if any(v is not None for v in (acc.get("values") or {}).values())
    )
    if accounts_total > 0 and accounts_with_values == 0:
        raise ValueError(
            f"build_excel: {accounts_total} accounts present but all value "
            "dicts are empty — likely a key-type mismatch between consolidate "
            "(int keys) and the JSON-roundtripped payload (string keys)."
        )

    # _apply_review routed open-question accounts into named groups
    groups = _apply_review_answers(groups, questions, review_answers)

    # "Verminderung" wirkt semantisch mindernd im Jahresergebnis. Die Behandlung
    # passiert direkt in der JÜ-Formel (immer subtrahieren), nicht über type.
    groups = _reclassify_bestandsveraenderung(groups)

    # Bilanzgewinn-Gruppen aus dem GuV-Body raustrennen, separat im
    # Bilanzgewinn-Block rendern.
    guv_groups = [g for g in groups
                   if g.get("gkv_section") not in BILANZGEWINN_SECTIONS]
    bilanzgewinn_groups = [g for g in groups
                            if g.get("gkv_section") in BILANZGEWINN_SECTIONS]
    groups = guv_groups

    # Sign-convention aus ECHTEN Aufwands-Werten ableiten, statt Claude zu vertrauen.
    columns = _infer_sign_conventions(columns, groups)

    # Header row
    ws.cell(row=1, column=1, value="Konto").font = BOLD
    ws.cell(row=1, column=2, value="Bezeichnung").font = BOLD
    for col_idx, col in enumerate(columns):
        c = ws.cell(row=1, column=3 + col_idx, value=col["label"])
        c.font = BOLD

    # Pre-compute: which groups are parents of sub-groups
    children_by_parent: dict[str, list[str]] = {}
    for g in groups:
        parent = g.get("sub_group_of")
        if parent:
            children_by_parent.setdefault(parent, []).append(g["name"])

    # Rows
    group_sum_rows: dict[str, int] = {}  # group_name -> row
    row_cursor = 2

    for g in groups:
        # Summen-Zeile zuerst
        sum_row = row_cursor
        group_sum_rows[g["name"]] = sum_row
        is_sub = g.get("sub_group_of") is not None
        label = ("  " if is_sub else "") + g["name"]
        cell_label = ws.cell(row=sum_row, column=2, value=label)
        cell_label.font = BOLD if not is_sub else BOLD_LIGHT

        has_children = g["name"] in children_by_parent and not g.get("accounts")

        # Accounts
        detail_start = row_cursor + 1
        detail_end = detail_start - 1
        for acc in g.get("accounts", []):
            row_cursor += 1
            ws.cell(row=row_cursor, column=1, value=acc.get("konto_nr") or "")
            ws.cell(row=row_cursor, column=2, value=f"  {acc.get('bezeichnung', '')}")
            for col_idx in range(len(columns)):
                v = acc.get("values", {}).get(col_idx)
                c = ws.cell(row=row_cursor, column=3 + col_idx, value=v)
                c.number_format = EUR_FORMAT
                if acc.get("confidence") == "low":
                    c.fill = YELLOW
            detail_end = row_cursor

        # Summen-Zeile befüllen
        for col_idx, col in enumerate(columns):
            target_col = 3 + col_idx
            col_letter = get_column_letter(target_col)
            has_details = detail_end >= detail_start
            if has_details:
                # Gruppe hat Konten — SUM-Formel (gilt für JA *und* BWA,
                # seit BWA auch Einzelkonten liefert)
                formula = f"=SUM({col_letter}{detail_start}:{col_letter}{detail_end})"
                c = ws.cell(row=sum_row, column=target_col, value=formula)
            elif col["kind"] == "bwa":
                # BWA ohne Konten für diese Gruppe → direkter Wert aus column_sums
                bwa_val = g.get("column_sums", {}).get(col_idx)
                c = ws.cell(row=sum_row, column=target_col, value=bwa_val)
            elif has_children:
                # Parent-Gruppe → SUM über Sub-Summen (Pass 2)
                c = ws.cell(row=sum_row, column=target_col, value=None)
            else:
                c = ws.cell(row=sum_row, column=target_col, value=0)
            c.number_format = EUR_FORMAT
            c.font = BOLD if not is_sub else BOLD_LIGHT

        row_cursor += 1

    # Pass 2: Parent-Gruppen-Summen über Sub-Summen. Wenn der Parent bereits
    # eine SUM(...)-Formel über seine eigenen Konten hat, ergänzen wir die
    # Sub-Refs daran -- so funktioniert auch der Fall "Top-Level mit Konten
    # UND Sub-Gruppen" (zB Sonst. betr. Aufw. mit direktem Forderungsverlust-
    # Konto und mehreren Sub-Kategorien).
    for parent_name, children_names in children_by_parent.items():
        parent_row = group_sum_rows.get(parent_name)
        if parent_row is None:
            continue
        for col_idx, col in enumerate(columns):
            target_col = 3 + col_idx
            col_letter = get_column_letter(target_col)
            current = ws.cell(row=parent_row, column=target_col).value
            child_refs = [f"{col_letter}{group_sum_rows[cn]}"
                          for cn in children_names if cn in group_sum_rows]
            if not child_refs:
                continue
            if isinstance(current, str) and current.startswith("=SUM("):
                # Parent hat eigene Konten-Summe -> Sub-Refs anhaengen
                formula = current + "+" + "+".join(child_refs)
            elif current is None:
                formula = "=" + "+".join(child_refs)
            else:
                # BWA-Direct-Wert oder 0 -> nicht ueberschreiben
                continue
            c = ws.cell(row=parent_row, column=target_col, value=formula)
            c.number_format = EUR_FORMAT
            c.font = BOLD

    # Leerzeile
    row_cursor += 1

    # Jahresergebnis-Zeile: Summe aller Top-Level-Gruppen (sub_group_of is None)
    # unter Berücksichtigung von Gruppen-Typ und sign_convention der jeweiligen Spalte.
    # Bei EÜR steht hier der Endwert-Begriff der Extraktion (z.B. "Steuerlicher Gewinn").
    je_row = row_cursor
    ws.cell(row=je_row, column=2, value=formula_label).font = BOLD
    formula_groups = _endwert_groups(groups)
    for col_idx, col in enumerate(columns):
        target_col = 3 + col_idx
        col_letter = get_column_letter(target_col)
        conv = col.get("sign_convention", "expenses_negative")
        # Wenn die Spalte echte Konten-Daten hat (egal in welcher Ebene), sind
        # reine Aggregat-Gruppen (accounts=0, keine Subs, nur column_sum)
        # redundante Sichten derselben Werte (typisch BWA-Aggregate die
        # JA-Konten zusammenfassen) → ueberspringen, sonst Doppelzaehlung.
        col_has_account_data = any(
            (acc.get("values") or {}).get(col_idx) is not None
            for g in groups for acc in (g.get("accounts") or [])
        )
        parts_plus: list[int] = []
        parts_minus: list[int] = []
        for g in formula_groups:
            is_bare_top_level = (
                g.get("sub_group_of") is None
                and not g.get("accounts")
                and not any(s.get("sub_group_of") == g["name"] for s in groups)
            )
            if col_has_account_data and is_bare_top_level:
                continue  # redundante BWA-Aggregat-Sicht
            sum_r = group_sum_rows.get(g["name"])
            if sum_r is None:
                continue
            role = _resolve_role(g)
            if conv == "expenses_negative":
                # Alles addieren — Aufwände sind eh negativ
                parts_plus.append(sum_r)
            else:
                if role == "ertrag":
                    parts_plus.append(sum_r)
                elif role in ("aufwand", "steuer"):
                    parts_minus.append(sum_r)
                else:
                    parts_plus.append(sum_r)
        if not parts_plus and not parts_minus:
            formula = "=0"
        else:
            plus = "+".join(f"{col_letter}{r}" for r in parts_plus) or "0"
            minus = ""
            if parts_minus:
                minus = "-" + "-".join(f"{col_letter}{r}" for r in parts_minus)
            formula = f"={plus}{minus}"
        c = ws.cell(row=je_row, column=target_col, value=formula)
        c.number_format = EUR_FORMAT
        c.font = BOLD

    # Bilanzgewinn-Block (Gewinnvortrag, Ausschuettung, Bilanzgewinn-Formel)
    bg_sum_rows: dict[str, int] = {}  # gkv_section -> row
    if bilanzgewinn_groups:
        row_cursor += 1
        ws.cell(row=row_cursor, column=2,
                value="--- Bilanzgewinn-Rechnung ---").font = BOLD_LIGHT
        for g in bilanzgewinn_groups:
            row_cursor += 1
            sum_row = row_cursor
            bg_sum_rows[g.get("gkv_section", "")] = sum_row
            ws.cell(row=sum_row, column=2, value=g["name"]).font = BOLD_LIGHT
            detail_start = row_cursor + 1
            detail_end = detail_start - 1
            for acc in g.get("accounts", []):
                row_cursor += 1
                ws.cell(row=row_cursor, column=1, value=acc.get("konto_nr") or "")
                ws.cell(row=row_cursor, column=2,
                        value=f"  {acc.get('bezeichnung', '')}")
                for col_idx in range(len(columns)):
                    v = acc.get("values", {}).get(col_idx)
                    c = ws.cell(row=row_cursor, column=3 + col_idx, value=v)
                    c.number_format = EUR_FORMAT
                detail_end = row_cursor
            for col_idx in range(len(columns)):
                target_col = 3 + col_idx
                col_letter = get_column_letter(target_col)
                if detail_end >= detail_start:
                    formula = (f"=SUM({col_letter}{detail_start}:"
                                f"{col_letter}{detail_end})")
                    c = ws.cell(row=sum_row, column=target_col, value=formula)
                else:
                    c = ws.cell(row=sum_row, column=target_col, value=0)
                c.number_format = EUR_FORMAT
                c.font = BOLD_LIGHT

        # Bilanzgewinn-Formel: JÜ + Gewinnvortrag - Ausschuettung
        # (- Verlustvortrag wenn vorhanden, + bilanzgewinn-Sektion bedeutet
        # die PDF gibt einen vorberechneten Wert vor — den ueberschreiben wir
        # nicht, sondern stellen daneben unsere Formel)
        row_cursor += 1
        bg_row = row_cursor
        ws.cell(row=bg_row, column=2,
                value="Bilanzgewinn (Formel)").font = BOLD
        gv_row = bg_sum_rows.get("gewinnvortrag")
        au_row = bg_sum_rows.get("ausschuettung")
        for col_idx in range(len(columns)):
            target_col = 3 + col_idx
            col_letter = get_column_letter(target_col)
            parts = [f"{col_letter}{je_row}"]
            if gv_row:
                parts.append(f"+{col_letter}{gv_row}")
            if au_row:
                parts.append(f"-{col_letter}{au_row}")
            formula = "=" + "".join(parts)
            c = ws.cell(row=bg_row, column=target_col, value=formula)
            c.number_format = EUR_FORMAT
            c.font = BOLD

    # Plausibilitaets-Anker: PDF-Endwert + Differenz-Zeile
    if pdf_jue_per_column:
        row_cursor += 1
        pdf_jue_row = row_cursor
        ws.cell(row=pdf_jue_row, column=2,
                value=f"PDF-{pdf_anker_label}").font = BOLD_LIGHT
        for col_idx in range(len(columns)):
            v = pdf_jue_per_column.get(col_idx)
            c = ws.cell(row=pdf_jue_row, column=3 + col_idx, value=v)
            c.number_format = EUR_FORMAT
            c.font = BOLD_LIGHT
        row_cursor += 1
        diff_row = row_cursor
        ws.cell(row=diff_row, column=2,
                value="Differenz Excel ↔ PDF").font = BOLD_LIGHT
        for col_idx in range(len(columns)):
            target_col = 3 + col_idx
            col_letter = get_column_letter(target_col)
            if pdf_jue_per_column.get(col_idx) is None:
                continue
            formula = f"={col_letter}{je_row}-{col_letter}{pdf_jue_row}"
            c = ws.cell(row=diff_row, column=target_col, value=formula)
            c.number_format = EUR_FORMAT
            c.font = BOLD_LIGHT

        # Build-Zeit-Cross-Check: Excel-JUE numerisch berechnen und gegen PDF-JUE
        # vergleichen. Diff > 1 ct → ValueError, Job geht auf FAILED. Damit kein
        # silently falsches Excel ausgeliefert wird.
        excel_jue = _compute_excel_jue_per_column(groups, columns)
        jue_errors = []
        for col_idx, col in enumerate(columns):
            pdf_v = pdf_jue_per_column.get(col_idx)
            excel_v = excel_jue.get(col_idx)
            if pdf_v is None or excel_v is None:
                continue
            if abs(excel_v - pdf_v) > 0.01:
                jue_errors.append(
                    f"{col.get('label')}: PDF-JÜ {pdf_v:,.2f} ≠ Excel-JÜ "
                    f"{excel_v:,.2f} (Diff {excel_v - pdf_v:+,.2f})"
                )
        if jue_errors:
            raise ValueError(
                f"Excel-{formula_label} stimmt nicht mit PDF-{pdf_anker_label} "
                "überein. Mögliche Ursachen: fehlende/falsch extrahierte Konten, "
                "Vorzeichen-Probleme, Cross-Year-Routing. Bitte PDF prüfen oder "
                "Job neu starten.\nBetroffen:\n  - " + "\n  - ".join(jue_errors)
            )

    # Spaltenbreiten
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 55
    for idx in range(len(columns)):
        ws.column_dimensions[get_column_letter(3 + idx)].width = 14

    # Fragen-Sheet — nur anlegen wenn echte User-Entscheidungen offen sind
    # (z.B. unmatched_account). Reine Audit-Mismatches (previous_year, group_sum,
    # jue) werden intern aufgelöst bzw. failen den Job — sie kommen hier nicht
    # mehr an.
    if questions:
        fragen = wb.create_sheet("Fragen")
        fragen.append(["Thema", "Details"])
        for q in questions:
            details = _format_question(q)
            fragen.append([q.get("type", ""), details])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _reclassify_bestandsveraenderung(groups: list[dict]) -> list[dict]:
    """Robustifiziert die GKV-Klassifikation per Name-Match (Defense-in-Depth
    falls Claude `gkv_section` nicht oder falsch setzt).

    GKV §275 Pos 2: Bestandsveränderung. Die Werte sind beim Konsolidieren
    bereits normalisiert: positiv = Erhöhung (+JÜ), negativ = Verminderung
    (-JÜ). Daher wird die Gruppe hier IMMER als Ertrag-Position eingestuft
    und in der JÜ-Formel addiert — das Vorzeichen entscheidet die Wirkung.

    'Gewinnvortrag', 'Verlustvortrag', 'Bilanzgewinn', 'Bilanzverlust',
    'Ausschüttung' sind Eigenkapital-Bewegungen (Bilanzgewinn-Rechnung), NICHT
    Teil der GuV. Werden hier zur passenden gkv_section-Sektion umgemappt,
    damit der Builder sie in den Bilanzgewinn-Block schiebt.
    """
    out = []
    for g in groups:
        name_lc = (g.get("name") or "").lower()
        if "gewinnvortrag" in name_lc or "verlustvortrag" in name_lc:
            out.append({**g, "gkv_section": "gewinnvortrag"})
        elif "ausschüttung" in name_lc or "ausschuettung" in name_lc:
            out.append({**g, "gkv_section": "ausschuettung"})
        elif ("bilanzgewinn" in name_lc or "bilanzverlust" in name_lc):
            out.append({**g, "gkv_section": "bilanzgewinn"})
        elif (("verminderung" in name_lc or "erhöhung" in name_lc
                or "bestandsveränderung" in name_lc)
               and "bestand" in name_lc):
            out.append({**g, "type": "ertrag",
                        "gkv_section": g.get("gkv_section") or "bestandsveraenderung"})
        else:
            out.append(g)
    return out


def _infer_sign_conventions(columns: list[dict], groups: list[dict]) -> list[dict]:
    """Leite sign_convention pro Spalte aus den echten Aufwand/Steuer-Werten ab.
    Mehrheit negativ → expenses_negative (einfache Summen-Formel fürs Jahresergebnis).
    Mehrheit positiv → expenses_positive (Aufwände müssen subtrahiert werden)."""
    out = []
    for col_idx, col in enumerate(columns):
        new_col = dict(col)
        samples: list[float] = []
        for g in groups:
            if g.get("type") not in ("aufwand", "steuer"):
                continue
            for acc in g.get("accounts", []):
                v = acc.get("values", {}).get(col_idx)
                if v is None:
                    continue
                try:
                    v = float(v)
                except (TypeError, ValueError):
                    continue
                if abs(v) < 0.01:
                    continue
                samples.append(v)
        if samples:
            # Summe statt Anzahl: einzelne Skonto-Korrekturen (viele kleine
            # negative Werte) duerfen die grossen Hauptaufwand-Werte nicht
            # ueberstimmen.
            total = sum(samples)
            new_col["sign_convention"] = (
                "expenses_negative" if total < 0 else "expenses_positive"
            )
        out.append(new_col)
    return out


def _apply_review_answers(groups: list[dict], questions: list[dict],
                          review_answers: dict) -> list[dict]:
    """Route accounts from open_questions into the user-chosen group."""
    if not review_answers:
        return groups

    # Index für schnelles Gruppen-Lookup
    groups_by_name = {g["name"]: g for g in groups}

    for q in questions:
        if q.get("type") != "unmatched_account":
            continue
        konto_nr = q.get("konto_nr")
        if not konto_nr or konto_nr not in review_answers:
            continue
        target_name = review_answers[konto_nr]
        target = groups_by_name.get(target_name)
        if target is None:
            continue
        # Wir haben hier nur einen Einzelwert (betrag_gj). Column-Index 0 annehmen
        # wenn Spalten bekannt sind — aber hier fehlt der Kontext. Review-Accounts
        # werden aktuell nur mit betrag_gj aus dem Dokument der jüngsten JA geroutet.
        # TODO: falls Multi-Jahr-Review nötig, müsste q eine column-Liste mitbringen.
        acc_key = f"reviewed:{konto_nr}"
        acc = {
            "konto_nr": konto_nr,
            "bezeichnung": q.get("bezeichnung", ""),
            "values": {0: q.get("betrag_gj")},
            "confidence": "reviewed",
        }
        target.setdefault("accounts", []).append(acc)

    return groups


def _compute_excel_jue_per_column(groups: list[dict],
                                    columns: list[dict]) -> dict[int, float]:
    """Berechnet die numerische JUE pro Spalte parallel zur Excel-Formel,
    damit ein Build-Zeit-Cross-Check gegen pdf_jue moeglich ist.
    Spiegelt die Logik der JE-Formel im Hauptbuilder (siehe _endwert_groups)."""
    out: dict[int, float] = {}
    formula_groups = _endwert_groups(groups)
    for col_idx, col in enumerate(columns):
        conv = col.get("sign_convention", "expenses_negative")
        col_has_account_data = any(
            (acc.get("values") or {}).get(col_idx) is not None
            for g in groups for acc in (g.get("accounts") or [])
        )
        total = 0.0
        for g in formula_groups:
            if g.get("gkv_section") in BILANZGEWINN_SECTIONS:
                continue
            is_bare_top_level = (
                g.get("sub_group_of") is None
                and not g.get("accounts")
                and not any(s.get("sub_group_of") == g["name"] for s in groups)
            )
            if col_has_account_data and is_bare_top_level:
                continue
            # Gruppen-Summe: eigene Konten + Konten der Sub-Gruppen wenn ein
            # Top-Level mit eigenen accounts ist (HGB-Pattern). Sub-Gruppen
            # einzeln im _endwert_groups-Output haben nur ihre eigenen Konten.
            grp_sum = 0.0
            for acc in g.get("accounts", []):
                v = acc.get("values", {}).get(col_idx)
                if isinstance(v, (int, float)):
                    grp_sum += v
            if g.get("sub_group_of") is None:
                for sub in groups:
                    if sub.get("sub_group_of") == g["name"]:
                        for acc in sub.get("accounts", []):
                            v = acc.get("values", {}).get(col_idx)
                            if isinstance(v, (int, float)):
                                grp_sum += v
            # BWA-Aggregat: Wert aus column_sums uebernehmen.
            if grp_sum == 0.0 and not g.get("accounts"):
                cs = (g.get("column_sums") or {}).get(col_idx)
                if isinstance(cs, (int, float)):
                    grp_sum = cs
            role = _resolve_role(g)
            if conv == "expenses_negative":
                total += grp_sum
            else:
                if role == "ertrag":
                    total += grp_sum
                elif role in ("aufwand", "steuer"):
                    total -= grp_sum
                else:
                    total += grp_sum
        out[col_idx] = total
    return out


def _format_question(q: dict) -> str:
    t = q.get("type", "")
    if t == "previous_year_mismatch":
        return (f"Gruppe {q.get('group')} · Konto {q.get('konto_nr')} · "
                f"Jahr {q.get('year')}: "
                f"PDF {q.get('from_doc_year')} sagt {q.get('pdf_says')}, "
                f"eigene Extraktion {q.get('own_value')}")
    if t == "group_sum_mismatch":
        return (f"Gruppe {q.get('group')} · Jahr {q.get('year')}: "
                f"PDF-Summe {q.get('pdf_says')} ≠ Konten-Summe {q.get('accounts_sum')}")
    if t == "unmatched_account":
        return (f"Konto {q.get('konto_nr') or '(ohne Nr)'} · "
                f"'{q.get('bezeichnung')}' · Jahr {q.get('year')} · "
                f"Betrag {q.get('betrag_gj')} — {q.get('hint', '')}")
    if t == "jue_excel_vs_pdf_mismatch":
        return (f"Spalte {q.get('column_label')}: "
                f"PDF-JÜ {q.get('pdf_says')} ≠ Excel-JÜ {q.get('excel_says')}")
    if t == "pdf_jue_previous_year_mismatch":
        return (f"Vorjahres-JÜ {q.get('year')}: "
                f"eigene PDF sagt {q.get('own_value')}, "
                f"PDF {q.get('from_doc_year')} sagt {q.get('pdf_says')}")
    return str(q)
