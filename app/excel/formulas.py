from openpyxl.utils import get_column_letter


def cell(col_idx: int, row: int) -> str:
    """A1-style reference. col_idx is 0-based."""
    return f"{get_column_letter(col_idx + 1)}{row}"


def sum_range(col_idx: int, first_row: int, last_row: int) -> str:
    col = get_column_letter(col_idx + 1)
    return f"=SUM({col}{first_row}:{col}{last_row})"


def multi_add(col_idx: int, add_rows: list[int], subtract_rows: list[int] | None = None) -> str:
    col = get_column_letter(col_idx + 1)
    parts = [f"{col}{r}" for r in add_rows]
    formula = "=" + "+".join(parts) if parts else "=0"
    for r in (subtract_rows or []):
        formula += f"-{col}{r}"
    return formula


def ratio(numerator_row: int, denominator_row: int, col_idx: int) -> str:
    """Ratio with division-by-zero guard. Produces percentage-safe formula."""
    col = get_column_letter(col_idx + 1)
    return f'=IFERROR({col}{numerator_row}/{col}{denominator_row},"")'


def safe_ref(col_idx: int, row: int | None) -> str:
    """Reference a cell if row is given, otherwise literal 0.
    Fix 2B: avoid falling back to Row 1 (header row) which produces #VALUE!."""
    if row is None:
        return "0"
    return cell(col_idx, row)
